# Builds report-17.07.26.xlsx — summary of the day's work: company workspace endpoints/layout/list,
# dashboard bucket toggle, trips-history tracker, cargo-create fixes, chat delete/pin, admin drivers
# filter panel + edit gate + status-gated dropdown. Run: python build-xlsx.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'report-17.07.26.xlsx')

BLUE = '1639C8'; DARK = '0F1D4F'; GREEN = '00B87A'; OKG = '0F9D58'
LINE = 'E6E9F0'; MUTEDBG = 'FAFBFE'
thin = Side(style='thin', color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)
top = Alignment(vertical='top')

wb = Workbook()

# ── Sheet 1: Changes ────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Changes'
ws.sheet_view.showGridLines = False

title = ws.cell(row=1, column=1, value='Sarbon — Company / Dashboard / Trips / Chat / Admin-drivers · 17.07.2026')
title.font = Font(bold=True, size=15, color=DARK)
ws.cell(row=2, column=1,
        value='Company workspace: 13 live endpoints wired, fleet hidden for SHIPPER, redesigned list + UI polish. Dashboard bar labels + day/week/month. Trips history tracker. cargo-create temp + ASAP. Chat delete (self-only) + pins. Admin drivers server-side filters + edit gate + status-gated dropdown. 84 files / 5 commits.'
        ).font = Font(size=10, color='5B6477', italic=True)
ws.merge_cells('A2:F2')

headers = ['#', 'Type', 'Area', 'Change', 'Endpoint / detail', 'Screenshot']
hr = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=BLUE)
    cell.alignment = Alignment(vertical='center', horizontal='left')
    cell.border = border

rows = [
    ['1', 'FEATURE', 'Company / Workspace',
     'Wired 13 live-spec endpoints (own-cargo offers, trip driver-assign/cancel/rating, trip payments, finance receivables/export/invoice, dashboard, quota, cargo lifecycle, expense edit, employee search); pure companyPayments/Dashboard/CargoActions.ts + tests',
     '/companies/:id/{offers,trips,finance,dashboard,quota,cargo}', '20-company-overview, 21-company-finance'],
    ['2', 'FIX', 'Company / Type gating',
     'Fleet (auto-park) group hidden for SHIPPER via forbiddenForTypes (Owner perms fail open, so type-gate not permission); ?tab=vehicles deep link falls back to overview instead of 403',
     'forbiddenForTypes: [SHIPPER]', '25-company-shipper-no-fleet'],
    ['3', 'FEATURE', 'Company / List',
     'Redesigned /company list into a responsive card grid with a freight-type accent (carrier/blue, shipper/amber, broker/teal) + avatar + status hint',
     'GET /v1/auth/companies', '24-company-list-cards'],
    ['4', 'FIX', 'Company / UI polish',
     'Scroll-to-top on section change (reduced-motion aware); full-width layout aligned under the logo; Trucks/Trailers segmented icons moved into the label (were misaligned via AntD option.icon)',
     'CompanyDetailPage, VehiclesSection', '22-company-vehicles'],
    ['5', 'FEATURE', 'Dashboard',
     'Bar value labels (recharts LabelList, zeros hidden) + day/week/month bucket toggle re-aggregating client-side (can only roll up from the API granularity); default view flipped to bars',
     'DashboardPage ActivityChartCard', '06-dashboard-activity-chart'],
    ['6', 'FIX', 'Trips / History',
     'Always-visible Event/Status eyebrow labels (were hover-only); each step shows its number (last no longer a tick, cancelled not an X); completion tick / cancel X moved to the card header; duplicate terminal date suppressed',
     'GET /v1/dispatchers/trips/history', '08-trips-history'],
    ['7', 'FIX', 'Cargo-create',
     'temp_min/temp_max gated on isReeferTrailerType and stripped from the payload for non-reefer trailers (was backend 400); Step 2 ASAP checkbox now clears the delivery-time required error (validator reads asap fresh + revalidates)',
     'buildCargoUpsertPayload, Step2Route/Step3Transport', '09-cargo-create'],
    ['8', 'FEATURE', 'Chat',
     'Delete conversation hides it for the CURRENT USER only (peer keeps it; new message re-surfaces; 404 = idempotent already-hidden); front-end conversation pins persisted per user id (tolerant parse)',
     'DELETE /v1/chat/conversations/:id', '26-chat-list'],
    ['9', 'FEATURE', 'Admin / Drivers',
     'Server-side filter panel (search + work/account/driver/trailer selects + online/has_trips/has_offers tri-state + work_state + company/freelancer UUID), AND-combined with sort + pagination; graceful 400 keeps last-good rows; adminDriverFilters.ts (11 tests)',
     'GET /v1/admin/drivers', '01-admin-drivers-filters, 05-admin-drivers-ru'],
    ['10', 'FIX', 'Admin / Driver edit',
     'Account-status dropdown is status-aware (blocked->unblock only, active->block, inactive->activate); fields read-only unless moderation pending/rejected; account_status editable only for creator/seo/moderator; 409/403 localized',
     'availableModerationActions, driverEditGate.ts', '04-admin-driver-account-dropdown, 03-admin-driver-edit-drawer'],
]
type_fill = {'FIX': OKG, 'FEATURE': GREEN, 'UI': '7C3AED'}
r = hr + 1
for row in rows:
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = wrap if c in (4, 5, 6) else top
        cell.font = Font(size=10)
        if c == 2:
            cell.font = Font(size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=type_fill.get(val, '64748B'))
            cell.alignment = Alignment(vertical='center', horizontal='center')
    r += 1

widths = [5, 10, 22, 54, 40, 26]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for rr in range(hr + 1, r):
    ws.row_dimensions[rr].height = 66

# ── Sheet 2: Verification ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Verification')
ws2.sheet_view.showGridLines = False
ws2.cell(row=1, column=1, value='Verification gates').font = Font(bold=True, size=14, color=DARK)
gates = [
    ['Gate', 'Result'],
    ['tsc -b (full project type-check)', 'exit 0'],
    ['vitest run', '1879 / 1879 (173 files)'],
    ['eslint --max-warnings 0 (changed files)', 'clean'],
    ['Company workspace (qa/verify-company-workspace.mjs)', '18/18 sections, 0 raw i18n keys'],
    ['Locale parity', '15/15 (company.* / admin.* en/uz/ru + placeholders)'],
    ['Live run (staging: admin + yuk-menejer)', '14 screenshots, behaviour confirmed'],
    ['Adversarial review (3-4 lens Sonnet) on 15:58 / 17:19', 'findings addressed'],
    ['Files changed', '84 (5 commits)'],
]
gr = 3
for i, (a, b) in enumerate(gates):
    ca = ws2.cell(row=gr + i, column=1, value=a)
    cb = ws2.cell(row=gr + i, column=2, value=b)
    for cell in (ca, cb):
        cell.border = border
        cell.alignment = top
    if i == 0:
        for cell in (ca, cb):
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=BLUE)
    else:
        ca.font = Font(size=10)
        cb.font = Font(size=10, bold=True, color=OKG)
ws2.column_dimensions['A'].width = 52
ws2.column_dimensions['B'].width = 40

wb.save(OUT)
print('wrote', OUT)
