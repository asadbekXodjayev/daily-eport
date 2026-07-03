# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:\Users\hp\Desktop\report\daily-report\03.07.26\report-03.07.26.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1639C8")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="16181D")
SUB_FONT = Font(bold=False, size=10.5, color="5b6477")
CELL_FONT = Font(bold=False, size=10.5, color="16181D")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CTR = Alignment(wrap_text=True, vertical="center")
THIN = Side(style="thin", color="E6E9F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

# ---------- Sheet 1: Сводка ----------
ws = wb.active
ws.title = "Сводка"
ws.column_dimensions["A"].width = 50
ws.column_dimensions["B"].width = 3
ws.column_dimensions["C"].width = 34
ws["A1"] = "Отчёт о работе · 03.07.2026 · Sarbon Frontend"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Фокус: видеозвонок (WebRTC), видео-заметки, поиск адреса, админ командный центр, модерация, грузы"; ws["A2"].font = SUB_FONT
rows = [
    ("03.07.2026", "Дата"),
    ("sarbon-frontend  ·  React 19 · Vite · Ant Design 6", "Проект"),
    ("staging", "Ветка / окружение"),
    ("ba518948 · 9cde2a45 · f23c1fdc · 5a9235ab", "Коммиты за день"),
    ("9", "Областей изменений"),
    ("4", "Функций добавлено"),
    ("3", "Дефектов закрыто"),
    ("~15", "Новых файлов"),
    ("~7", "Новых наборов тестов"),
    ("15", "Локалей (~55 новых ключей)"),
    ("0", "Ошибок tsc -b"),
    ("0", "Предупреждений ESLint (затронутые файлы)"),
]
r = 4
for val, label in rows:
    ws.cell(row=r, column=1, value=val).font = CELL_FONT
    ws.cell(row=r, column=3, value=label).font = SUB_FONT
    r += 1

# ---------- Sheet 2: Задачи ----------
ws2 = wb.create_sheet("Задачи")
cols2 = [("A",5),("B",16),("C",34),("D",42),("E",46),("F",50),("G",38),("H",24)]
for c,w in cols2: ws2.column_dimensions[c].width = w
head2 = ["№","Тип","Область","Что было / причина","Корень / детали","Что сделано","Файлы","Статус"]
for i,h in enumerate(head2, start=1):
    c = ws2.cell(row=1, column=i, value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=WRAP_CTR; c.border=BORDER
tasks = [
    ["1","Функция","Видеозвонок (WebRTC) + демонстрация экрана",
     "Был только аудиозвонок — без видео, камеры, шаринга экрана и подстройки под слабую сеть",
     "Видео-трансивер сделан всегда sendrecv (даже в аудио), поэтому видео/шаринг включаются без пересоздания сессии; удалённое видео — по unmute трека; тип звонка audio|video|screen — UI-подсказка, медиа задаёт SDP",
     "Камера вкл/выкл, getDisplayMedia (шаринг экрана), локальное превью + удалённое видео, сворачивание/разворачивание панели, входящий «ответить с видео / только аудио»; webrtcTuning: кодек-предпочтение, ограничение битрейта/FPS, degradation preference",
     "GlobalAudioCallModal.tsx, webrtcTuning.ts (+тест), callTypes.ts, useCallStore.ts, ChatPage.tsx, ChatThreadPanel.tsx, useDispatcherSseSubscriptions.ts, locales ×15",
     "✅ Закоммичено f23c1fdc, 5a9235ab"],
    ["2","Функция","Видео-заметки в чате («кружочки»)",
     "Нельзя было записать короткое видеосообщение прямо из композера чата",
     "Запись гейтится доступом к камере/мик. и защищённым контекстом (HTTPS); все состояния отказа обработаны отдельно",
     "Запрос доступа, выбор/переключение камеры, лимиты длительности, выбор MIME/битрейта, оптимистичный «uploading»-пузырёк, воспроизведение с кольцом-прогрессом",
     "useVideoNoteRecorder.ts, chatVideoNoteRecorder.ts (+тест), ChatVideoNoteRecorder.tsx, ChatVideoNoteMessage.tsx, ChatVideoNoteRing.tsx, ChatComposerPanel.tsx, MessageBubble.tsx, ConversationItem.tsx",
     "✅ Закоммичено 9cde2a45, f23c1fdc"],
    ["3","Функция","Груз — поиск адреса маршрута (Yandex)",
     "Точку маршрута выбирали в основном на карте; текстового автодополнения адреса не было",
     "Геокодер Yandex; запрос с 3-го символа, дебаунс 400 мс, до 7 результатов; резолв в мультиязычные поля адреса + город/таймзона",
     "Поле поиска адреса с автодополнением, выбор точки резолвится в address_uz/ru/en/zh + город/таймзона",
     "RoutePointAddressSearch.tsx, resolveYandexRoutePoint.ts (+тест), Step2Route.tsx",
     "✅ Закоммичено 9cde2a45"],
    ["4","Функция","Админ — командный центр (переработка)",
     "Единый глобальный диапазон дат; счётчики модерации только по грузам/водителям; графики в основном потоке",
     "Пофильтровые диапазоны дат заменяют один глобальный; пер-роль pending берётся из analytics/moderation (dashboard-тоталы не покрывают менеджеров)",
     "Плитки-сводка по ролям; DateRangeFilterDropdown + фильтр колонки по датам; графики композиции/рейсов — в модалку; счётчики «в модерации» по ролям (CM/DM)",
     "AdminCommandCenterPage.tsx, DateRangeFilterDropdown.tsx, tableFilters.tsx (+тест), AdminKpiTile.tsx, commandCenterChartData.ts",
     "✅ Закоммичено 9cde2a45"],
    ["5","Изменение","Админ — модерация водителей (KYC → допуск)",
     "Вкладка «Водители» модерировала KYC (личность), требовалось модерировать допуск (moderation_status), как у менеджеров",
     "Бэкенд отдаёт отдельную поверхность допуска: GET /admin/drivers/moderation + POST …/moderation/{approve,reject}",
     "Панель переписана по образцу менеджеров: очередь moderation_status=pending, Accept/Reject через выделенные эндпоинты, колонка moderation_status (KYC-колонка убрана)",
     "DriverModerationPanel.tsx, AdminModerationPage.tsx, locales ×15",
     "✅ Закоммичено 9cde2a45; ⚠ live-раунд на staging"],
    ["6","Функция","Гейт модерации диспетчера + модерация менеджеров",
     "Не было ограничения доступа для «pending»-диспетчеров; вкладку «Менеджеры» нужно перевести на выделенные эндпоинты",
     "moderation_status пока НЕ приходит в /v1/dispatchers/profile — гейт «fail-open» до появления поля на бэкенде",
     "Экран ограниченного доступа для pending/rejected (kill-switch в env); модерация менеджеров переведена на DISPATCHER_MODERATION_{approve,reject}",
     "ModerationGate.tsx, Account{UnderReview,Rejected}Screen.tsx, moderationGate.ts (+тест), guestPaths.ts, ProtectedRoute.tsx, ManagerModerationPanel.tsx, endpoints.ts",
     "✅ Закоммичено ba518948; ⚠ бэкенд-зависимость"],
    ["7","Фикс (bug)","Груз — вкладка «Все» показывала не все статусы",
     "Вкладка «Все» отдавала только грузы «Searching (all)» — модерация/в работе/завершённые/архив пропадали",
     "TAB_TO_API_STATUS.all был undefined → параметр status опускался, а бэкенд по умолчанию сужает выдачу до «поиска», а не «всё»",
     "Вкладка «Все» шлёт объединение статусов всех остальных вкладок (выводится программно); Excel-экспорт со вкладки «Все» тоже стал полным",
     "myCargosConstants.ts (+тест myCargosTabStatus)",
     "✅ Закоммичено 9cde2a45"],
    ["8","Фикс (bug)","Груз — редактирование (PUT → PATCH)",
     "Сохранение правок деактивированного груза падало с 400; бэкенд перевёл обновление груза с PUT на PATCH",
     "Method mismatch — FE слал PUT, бэкенд теперь принимает PATCH /api/cargo/{id}",
     "Мутация обновления груза переведена usePut → usePatch (byte-identical drop-in); деактивированные грузы снова сохраняются",
     "MyCargoEditForm.tsx, docs/backend/cargo-update-deactivated-400.md",
     "✅ Закоммичено 9cde2a45; бэкенд подтверждён live"],
    ["9","Фикс (bug)","Аудиозвонок — устойчивость к обрыву (перенос 02.07)",
     "При локальном сбое звонка модалка закрывалась, серверный звонок оставался ACTIVE → «залипание занят»",
     "cleanupCall('error') очищал всю сессию; isCallActive исключал error",
     "preserveSession сохраняет сессию на ошибках; модалка остаётся с кнопкой «Завершить»; 2 регрессии устранены в review",
     "GlobalAudioCallModal.tsx, ChatThreadPanel.tsx",
     "✅ Закоммичено ba518948"],
]
for ri,row in enumerate(tasks, start=2):
    for ci,val in enumerate(row, start=1):
        c = ws2.cell(row=ri, column=ci, value=val); c.font=CELL_FONT; c.alignment=WRAP_TOP; c.border=BORDER

# ---------- Sheet 3: Изменённые файлы ----------
ws3 = wb.create_sheet("Изменённые файлы")
cols3 = [("A",60),("B",26),("C",64),("D",26)]
for c,w in cols3: ws3.column_dimensions[c].width = w
head3 = ["Файл","Тип изменения","Что сделано","Коммит"]
for i,h in enumerate(head3, start=1):
    c = ws3.cell(row=1, column=i, value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=WRAP_CTR; c.border=BORDER
files = [
    ["src/components/shared/GlobalAudioCallModal.tsx","Логика/UI (звонки)","Видеозвонок: камера, шаринг экрана, локальное/удалённое видео, сворачивание панели, входящий видео","f23c1fdc, 5a9235ab, ba518948"],
    ["src/utils/webrtcTuning.ts","Утилита (новый)","Тюнинг слабой сети: кодек-предпочтение, битрейт/FPS, degradation preference (+тест)","f23c1fdc"],
    ["src/types/callTypes.ts","Типы","CallType = audio | video | screen; call_type в payload/сторе","f23c1fdc"],
    ["src/stores/useCallStore.ts","Стор","callType в состоянии звонка","f23c1fdc"],
    ["src/hooks/useVideoNoteRecorder.ts","Хук (новый)","Машина состояний записи видео-заметки (камера/разрешения/лимиты)","9cde2a45, f23c1fdc"],
    ["src/pages/chat/utils/chatVideoNoteRecorder.ts","Утилита (новый)","MediaRecorder-хелперы: MIME, constraints, битрейт, лимиты, маппинг ошибок (+тест)","9cde2a45, f23c1fdc"],
    ["src/pages/chat/components/ChatVideoNoteRecorder.tsx","UI (новый)","Экран записи видео-заметки","9cde2a45, f23c1fdc"],
    ["src/pages/chat/components/ChatVideoNoteMessage.tsx","UI (новый)","Пузырёк видео-заметки + воспроизведение","9cde2a45, f23c1fdc"],
    ["src/pages/chat/components/ChatVideoNoteRing.tsx","UI (новый)","Кольцо-прогресс вокруг кружочка","9cde2a45"],
    ["src/pages/chat/ChatPage.tsx","UI/логика (чат)","startCall(audio|video); оптимистичный «uploading»-пузырёк видео-заметки","9cde2a45, f23c1fdc"],
    ["src/pages/chat/components/ChatComposerPanel.tsx","UI (чат)","Кнопка/поток записи видео-заметки в композере","9cde2a45, 5a9235ab"],
    ["src/pages/chat/components/ChatThreadPanel.tsx","UI/логика (чат)","Кнопки видеозвонка; callBusy","f23c1fdc, ba518948"],
    ["src/pages/cargos/cargo-create/components/RoutePointAddressSearch.tsx","UI (новый, груз)","Поиск адреса маршрута с автодополнением (Yandex geocode)","9cde2a45"],
    ["src/utils/resolveYandexRoutePoint.ts","Утилита (новый)","Резолв гео-объекта в мультиязычные поля адреса (+тест)","9cde2a45"],
    ["src/pages/cargos/cargo-create/components/Step2Route.tsx","UI (груз)","Интеграция поиска адреса в шаг маршрута","9cde2a45"],
    ["src/pages/admin/command-center/AdminCommandCenterPage.tsx","UI (админ)","Плитки по ролям, пофильтровые даты, счётчики модерации по ролям","9cde2a45"],
    ["src/components/shared/admin/DateRangeFilterDropdown.tsx","UI (новый, админ)","Дропдаун диапазона дат для фильтра колонки","9cde2a45"],
    ["src/components/shared/admin/tableFilters.tsx","Логика (админ)","Фильтр колонки по диапазону дат (+тест)","9cde2a45"],
    ["src/pages/admin/moderation/components/DriverModerationPanel.tsx","UI (админ)","KYC → допуск: очередь moderation_status, approve/reject","9cde2a45"],
    ["src/components/shared/moderation/ModerationGate.tsx","UI/логика (новый)","Гейт: экран ограниченного доступа для pending/rejected","ba518948"],
    ["src/utils/moderationGate.ts","Логика (новый)","Предикат moderationGateDecision (fail-open) (+тест)","ba518948"],
    ["src/router/guestPaths.ts","Логика (новый)","Общая isGuestAllowedPath для ProtectedRoute и ModerationGate","ba518948"],
    ["src/pages/cargos/my-cargos/utils/myCargosConstants.ts","Логика (груз)","Вкладка «Все» = объединение статусов всех вкладок (+тест)","9cde2a45"],
    ["src/pages/cargos/my-cargo-detail/components/MyCargoEditForm.tsx","Логика (груз)","usePut → usePatch для обновления груза","9cde2a45"],
    ["public/locales/*.json (×15)","Локализация","~55 ключей: видеозвонок, видео-заметки, поиск адреса, модерация","все 4 коммита"],
    ["bug.md","Документация","4 записи 03.07 (10:45 · 14:00 · 14:30 · 15:05)","верх файла"],
]
for ri,row in enumerate(files, start=2):
    for ci,val in enumerate(row, start=1):
        c = ws3.cell(row=ri, column=ci, value=val); c.font=CELL_FONT; c.alignment=WRAP_TOP; c.border=BORDER

# ---------- Sheet 4: QA - Статус ----------
ws4 = wb.create_sheet("QA - Статус")
cols4 = [("A",44),("B",30),("C",66)]
for c,w in cols4: ws4.column_dimensions[c].width = w
head4 = ["Проверка","Результат","Примечание"]
for i,h in enumerate(head4, start=1):
    c = ws4.cell(row=1, column=i, value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=WRAP_CTR; c.border=BORDER
qa = [
    ["tsc -b (CI-ворота)","✅ 0 ошибок","Полная сборка проекта — единственная воротная проверка CI"],
    ["eslint --max-warnings 0","✅ 0 предупреждений","На затронутых файлах (pre-commit)"],
    ["Новые наборы тестов","✅ добавлены","webrtcTuning · chatVideoNoteRecorder · resolveYandexRoutePoint · dateRangeColumnFilter · myCargosTabStatus · moderationGate · dispatcherModeration"],
    ["Локализация","✅ 15 локалей","~55 новых ключей (видеозвонок, видео-заметки, поиск адреса, модерация)"],
    ["verify — видеозвонок / шаринг экрана","⚠ требует QA","P2P между 2 абонентами под авторизацией; в среде отчёта недоступно"],
    ["verify — видео-заметки","⚠ требует QA","Гейтится камерой/мик. и HTTPS; нужен клик на устройстве с камерой"],
    ["verify — поиск адреса маршрута","⚠ требует QA","Зависит от квоты/языка геокодера Yandex; проверить мультиязычный ответ"],
    ["Гейт модерации диспетчера","⚠ бэкенд-зависимость","moderation_status пока не в /v1/dispatchers/profile — «fail-open» до появления"],
    ["Модерация водителей/менеджеров (live)","⚠ на staging","approve/reject меняют реальные данные — не на проде"],
    ["Скриншоты админ-страниц","⚠ по запуску","capture-admin-command-center.mjs / capture-admin-driver-moderation.mjs на dev-сервере (staging)"],
]
for ri,row in enumerate(qa, start=2):
    for ci,val in enumerate(row, start=1):
        c = ws4.cell(row=ri, column=ci, value=val); c.font=CELL_FONT; c.alignment=WRAP_TOP; c.border=BORDER

for w in (ws2, ws3, ws4):
    w.freeze_panes = "A2"

wb.save(OUT)
print("wrote", OUT)
