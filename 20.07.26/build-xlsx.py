# Builds report-20.07.26.xlsx — summary of the day's work: company employee invitations
# investigation/fix, driver email display, admin trips/offers snake_case contract fix + cancel
# rename, city-name display, Excel export column-truncation + (re-investigated) locale fix,
# company-surface architecture cleanup, and two merged teammate PRs. Run: python build-xlsx.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'report-20.07.26.xlsx')

BLUE = '1639C8'; DARK = '0F1D4F'; GREEN = '00B87A'; OKG = '0F9D58'
LINE = 'E6E9F0'; MUTEDBG = 'FAFBFE'
thin = Side(style='thin', color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)
top = Alignment(vertical='top')

wb = Workbook()

# -- Sheet 1: Changes --------------------------------------------------------------------
ws = wb.active
ws.title = 'Changes'
ws.sheet_view.showGridLines = False

title = ws.cell(row=1, column=1, value='Sarbon - Company invitations / admin trips-offers / driver email / Excel / city names - 20.07.2026')
title.font = Font(bold=True, size=15, color=DARK)
ws.cell(row=2, column=1,
        value='Company employee invitations: /offers crash fix, accept/decline card+modal, notification redesign, backend blocker documented. Admin trips/offers snake_case contract fix + Cancel-trip rename. Driver detail email row. City-name display (comma + abbreviation). Excel export column-truncation fix + re-investigated locale fix (first attempt did not work). Company-surface architecture cleanup. Two merged teammate PRs. 98 files / 6 commits + working-tree.'
        ).font = Font(size=10, color='5B6477', italic=True)
ws.merge_cells('A2:F2')

headers = ['#', 'Type', 'Area', 'Change', 'Endpoint / detail', 'Screenshot']
hr = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=BLUE)
    cell.alignment = Alignment(vertical='center', horizontal='left')
    cell.border = border

rows = [
    ['1', 'FIX', 'Offers / Company invitations',
     'GET /v1/dispatchers/offers/all is polymorphic: invitation rows carry cargo:null + offer:null, but SentOfferListItem typed both non-nullable, so item.cargo.id crashed the whole /offers page (ErrorBoundary catch). Made nullable, split via isOfferListItemWithCargo / isCompanyInvitationOffer predicates, invitations render as a dedicated accept/decline card',
     'GET /v1/dispatchers/offers/all', '02-offers-invitation-card, 03-offers-invitation-modal'],
    ['2', 'FIX', 'Offers / Company invitations',
     'Root cause of "cannot accept": the invitee (CM/DM) is a separate person from the company owner/director and has no company account by design (confirmed via localStorage - no company token). The only endpoint that fits (dispatcher-scoped, already in production) is never given the invitation token it needs. Buttons enabled, but short-circuit client-side with an explanation instead of a guaranteed 401. Full diagnosis in company-invitation-offers-backend-ticket.md',
     'POST /v1/dispatchers/invitations/{accept,decline} (blocked - missing token)', '01-dispatcher-invitation-bell'],
    ['3', 'FEATURE', 'Drivers / Detail page',
     'SARB-DISP: GET /v1/dispatchers/drivers/{id} returns email, but the detail page never rendered it. Added an Email row next to phone, same readField/renderValue pattern, same "No data" fallback for empty values, no format validation',
     'GET /v1/dispatchers/drivers/{id}', '07-driver-detail-email'],
    ['4', 'FIX', 'Admin / Trips + Offers',
     'Backend switched /admin/trips and /admin/offers responses from PascalCase to snake_case without notice - every field silently read as "no data". Added a normalizeAdminRowKeys/pascalToSnake boundary normalizer covering both reads and the spots that wrote stale PascalCase keys back. Delete action renamed to "Cancel trip" (Trash2 -> Ban icon) since it is actually a soft-cancel on the backend',
     'GET /v1/admin/trips, /v1/admin/offers', '08-admin-trips, 09-admin-trip-detail-drawer, 11-admin-offers'],
    ['5', 'FIX', 'Cargo lists / City name',
     'Route origin/destination now display the city name up to the first comma everywhere (my-cargos, all-cargos, offers, admin) instead of either a bare city or the full address. Administrative words (oblast/viloyati/region) are abbreviated (obl./vil./reg.), not dropped; a leading city marker (g. Tashkent) is stripped',
     'summarizeCargoRoute / cityNameFromAddress', '04-all-cargos-city-name'],
    ['6', 'FIX', 'Excel export viewer',
     'Column truncation at Z fixed (col.len now computed explicitly instead of falling back to the widget default of 26) - verified. The "Chinese context menu" fix from the same morning commit turned out NOT to work: it called xSpreadsheet.locale(), which the bundled widget ignores for menu/toolbar text; the real switch is a DOM attribute (data-excel-viewer-lang) read fresh on every label lookup. Re-investigated and fixed properly, verified via a real click (not a manual override)',
     'excel-viewer package internals (xspreadsheet.js)', '05-excel-viewer-columns, 05b-excel-viewer-scrolled, 06-excel-viewer-context-menu'],
    ['7', 'REFACTOR', 'Architecture / Company surface',
     'arch-audit EVOLVE pass on the company workspace: extracted surface-neutral parsing helpers (toRecord/toRecordArray/getEnvelopeData/prettifyKey) out of adminHelpers.ts into recordParsing.ts (42 call sites re-exported, zero behavior change); moved duplicated employee-parsing helpers into companyEmployees.ts; removed a dead COMPANY_TYPES export. No visual changes',
     'src/utils/recordParsing.ts, companyEmployees.ts', '(no UI - structural only, gates re-run per step)'],
    ['8', 'FEATURE', 'Admin / Merged from teammate',
     'Two PRs merged into staging today (Abdumaliks, not this session\'s work): #297 splits the combined "Managers" moderation queue into Cargo managers / Driver managers tabs; #298 makes trip-detail Yuk ID / Haydovchi ID clickable links opening cargo/driver detail drawers (Offer ID stays copy-only - no backend endpoint for it), plus a fix for the same PascalCase cargo-detail bug on GET /v1/admin/cargo/{id}',
     '/admin/moderation?tab=cargo_managers|driver_managers', '12-admin-moderation-cargo-managers, 13-admin-moderation-driver-managers, 09b-admin-trip-system-info'],
]
type_fill = {'FIX': OKG, 'FEATURE': GREEN, 'UI': '7C3AED', 'REFACTOR': '64748B'}
r = hr + 1
for row in rows:
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = wrap if c in (4, 5, 6) else top
        cell.font = Font(size=10)
        if c == 2:
            cell.font = Font(size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=type_fill.get(val, '64748B'))
            cell.alignment = Alignment(vertical='center', horizontal='center')
    r += 1

widths = [5, 10, 24, 56, 40, 34]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for rr in range(hr + 1, r):
    ws.row_dimensions[rr].height = 84

# -- Sheet 2: Verification ---------------------------------------------------------------
ws2 = wb.create_sheet('Verification')
ws2.sheet_view.showGridLines = False
ws2.cell(row=1, column=1, value='Verification gates').font = Font(bold=True, size=14, color=DARK)
gates = [
    ['Gate', 'Result'],
    ['tsc -b (full project type-check)', 'exit 0'],
    ['vitest run --no-file-parallelism', '182/182 files, 2023/2023 tests'],
    ['eslint --max-warnings 0 (whole project)', 'clean'],
    ['pnpm build', 'exit 0'],
    ['Locale parity', '15/15 (driversPage.emailLabel, tripNotifications.invitationOffer.*, admin.pages.trips.cancel.*)'],
    ['Excel locale fix re-verified after first attempt failed', 'real click -> English, not a manual override'],
    ['Company-invitation accept/decline', 'BLOCKED on backend - see company-invitation-offers-backend-ticket.md'],
    ['bug.md', 'entry added for the re-investigated Excel fix'],
    ['Files touched', '98 (6 commits + working-tree, incl. 2 merged teammate PRs)'],
]
gr = 3
for i, (a, b) in enumerate(gates):
    ca = ws2.cell(row=gr + i, column=1, value=a)
    cb = ws2.cell(row=gr + i, column=2, value=b)
    for cell in (ca, cb):
        cell.border = border
        cell.alignment = wrap if i > 0 else top
    if i == 0:
        for cell in (ca, cb):
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=BLUE)
    else:
        ca.font = Font(size=10)
        cb.font = Font(size=10, bold=True, color=OKG if 'BLOCKED' not in b else 'E8590C')
ws2.column_dimensions['A'].width = 44
ws2.column_dimensions['B'].width = 64
for rr in range(gr, gr + len(gates)):
    ws2.row_dimensions[rr].height = 30

wb.save(OUT)
print('wrote', OUT)
