# Builds report-10.07.26.xlsx — PM summary of 10.07.2026 Sarbon Frontend work (3 sheets).
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = "1639C8"; SOFT = "EEF1FF"; INK = "16181D"
hdr_fill = PatternFill("solid", fgColor=BLUE)
sub_fill = PatternFill("solid", fgColor=SOFT)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
lbl_font = Font(bold=True, color=INK)
wrap = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="D9DEE8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = wrap; cell.border = border

def put(ws, r, c, v, font=None, fill=None, al=wrap):
    cell = ws.cell(row=r, column=c, value=v)
    if font: cell.font = font
    if fill: cell.fill = fill
    cell.alignment = al; cell.border = border
    return cell

# ---------- Sheet 1: Summary ----------
ws = wb.active; ws.title = "Сводка"
ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 82
put(ws, 1, 1, "Отчёт о работе · 10.07.2026 · Sarbon Frontend", hdr_font, hdr_fill)
put(ws, 1, 2, "Фокус: подготовка к релизу staging→main — редизайн админ Command Center, модерация, серверная пагинация, GPS «только свои», упрощённый вход", hdr_font, hdr_fill)
rows = [
    ("Дата", "10.07.2026"),
    ("Проект", "sarbon-frontend · React 19 · Vite · Ant Design 6"),
    ("Ветка", "staging → main (2 merge)"),
    ("Коммиты за день", "a953ba77 (merge), c46fa3f4, e2ae515b (merge)"),
    ("Главное", "Редизайн админ Command Center (KPI «Всего грузов» + единый Summary с counter-карточками)"),
    ("Command Center", "6 KPI · контролы тренда (диапазон/серии/скролл) · Summary: ~20 метрик со «семьёй» (value·рост%·avg·avg-рост%·prev) · акценты · переключатель на гистограмму · 3 визуальных фикса"),
    ("Модерация диспетч./водит.", "полный lifecycle: resubmit/активация/inactive/блок+каскад/разблок · ужатые действия · volcano-тег"),
    ("Пагинация админ-списков", "useAdminPaginatedList — server-paginated browse (50/стр, мгновенный рендер) vs fetch-all при фильтре/сортировке"),
    ("GPS (менеджер)", "карта только своих водителей (+ на активных рейсах); поиск по имени ИЛИ телефону (digits-only толерантно)"),
    ("Вход (release-prep)", "только телефон — идентификатор Telegram/Email скрыт (EMAIL_AUTH_ENABLED=false)"),
    ("AI-создание груза", "скрыто флагом AI_CARGO_CREATE_ENABLED=false (готово, вернём позже)"),
    ("Админ i18n-аудит", "захардкоженный английский в summary-вью заменён ключами локалей"),
    ("Тесты (pnpm test)", "989/989 (100 файлов)"),
    ("Ошибок tsc -b", "0"),
    ("Предупреждений ESLint", "0 (--max-warnings 0)"),
    ("Скриншоты-доказательства", "живой Playwright, авторизация: Admin, Cargo Manager +998994878460, Driver Manager +998998809935"),
]
for i, (k, v) in enumerate(rows, start=2):
    put(ws, i, 1, k, lbl_font, sub_fill); put(ws, i, 2, v)

# ---------- Sheet 2: Changes ----------
ws2 = wb.create_sheet("Изменения")
widths = [4, 14, 32, 46, 52, 34, 20]
for i, w in enumerate(widths, start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
headers = ["№", "Тип", "Область", "Что было / причина", "Что сделано", "Файлы", "Статус"]
for c, h in enumerate(headers, start=1): put(ws2, 1, c, h)
style_header(ws2, len(headers))
changes = [
    ("redesign", "Админ — редизайн Command Center",
     "Панель разбита на отдельные секции, метрики показывали лишь голое значение, нет KPI по грузам и контролов тренда.",
     "6-я KPI «Всего грузов»; тренд — диапазон дат/показ-скрытие серий/горизонтальный скролл; единый Summary: ~20 метрик counter-карточками со «семьёй» (value·рост%·avg·avg-рост%·prev), акцент и переключатель на гистограмму.",
     "command-center/** (summaryMetricConfig, AdminSummaryMetrics, AdminSummaryMetricCard, AdminMetricMiniTrend), charts/*, локали ×15", "✅ Закоммичено · test 989/989"),
    ("bug", "Command Center — 3 визуальных фикса",
     "Чипы-серии плохо читались; лишний вертикальный скроллбар графика; календарь диапазона за соседней секцией.",
     "Кастомные пилюли, тонированные цветом серии; overflow-y-hidden (скролл только по горизонтали); getPopupContainer=body для RangePicker.",
     "AdminCommandCenterPage.tsx, charts/AdminTrendBars.tsx", "✅ Закоммичено"),
    ("feature", "Админ — жизненный цикл модерации",
     "Действия строки разрозненны; не хватало полного цикла статусов доступа.",
     "resubmit из любого статуса, активация, inactive, блок с каскадом, разблок; ужатые действия строки, отдельная «Заблокировать», volcano-тег для заблокированных.",
     "AdminModerationActions.tsx, adminModerationRoutes.ts, AdminDispatchersPage.tsx, AdminDriversPage.tsx", "✅ Закоммичено"),
    ("perf", "Админ — серверная пагинация списков",
     "Excel-подобные списки грузили всё — десятки последовательных запросов блокировали первый рендер.",
     "useAdminPaginatedList: browse → server-paginated 50/стр (мгновенный первый рендер); filter/sort → fetch-all. Обе очереди смонтированы — порядок хуков стабилен.",
     "hooks/useAdminPaginatedList.ts (new), AdminDriversPage.tsx, AdminDispatchersPage.tsx, shared/table/Table.tsx", "✅ Закоммичено"),
    ("feature", "GPS — только свои водители + поиск по телефону",
     "Менеджер видел весь автопарк; фильтр по водителю не искал по телефону.",
     "Cargo/Driver-менеджеры видят только своих (свои + на активных рейсах); фильтр ищет по имени ИЛИ телефону, «как есть» и digits-only.",
     "GpsTrackingPage.tsx, gps-tracking/hooks/useGpsMarkersData.ts, GpsFilterSidebar.tsx, utils/driverOptionFilter.ts (new)", "✅ Закоммичено"),
    ("flag", "Вход — упрощён до телефона (release-prep)",
     "Переключатель идентификатора Telegram/Email не нужен, пока email-контур не готов e2e.",
     "EMAIL_AUTH_ENABLED=false прячет segmented-переключатель (остаются табы Пароль/Код над телефоном); бэкенд/хелперы email не тронуты. AI_CARGO_CREATE_ENABLED=false прячет AI-создание.",
     "constants/featureFlags.ts (new), LoginPage.tsx, MyCargosActionBar.tsx, RootRouter.tsx", "✅ Закоммичено"),
    ("bug", "Админ — i18n-аудит summary-вью",
     "Захардкоженный английский в summary-вью (users/communications/geo/metrics/moderation).",
     "Строки заменены ключами локалей; синхронизированы недостающие ключи.",
     "admin/{communications,metrics,moderation,users}/**/Admin*SummaryView.tsx, локали ×15", "✅ Закоммичено"),
]
for i, row in enumerate(changes, start=2):
    put(ws2, i, 1, i - 1)
    for c, v in enumerate(row, start=2):
        put(ws2, i, c, v)

# ---------- Sheet 3: Checks ----------
ws3 = wb.create_sheet("Проверки")
ws3.column_dimensions['A'].width = 40; ws3.column_dimensions['B'].width = 66
for c, h in enumerate(["Проверка", "Результат"], start=1): put(ws3, 1, c, h)
style_header(ws3, 2)
checks = [
    ("tsc -b (типы)", "0 ошибок"),
    ("eslint --max-warnings 0", "0 предупреждений (затронутые файлы)"),
    ("pnpm test", "989/989 (100 файлов), вкл. summaryMetricConfig.test.ts (8 кейсов)"),
    ("Визуальные баги Command Center", "3 закрыто и проверено на hot-reload"),
    ("Слияние в main", "2 merge-коммита (a953ba77 · e2ae515b)"),
    ("Серверная пагинация", "первый рендер больших списков мгновенный (тянется только видимая страница)"),
    ("Ролевое гейтирование", "подтверждено (driver-manager «Мои водители», cargo-manager «Менеджеры водителей»)"),
    ("Release-prep флаги", "EMAIL_AUTH_ENABLED=false · AI_CARGO_CREATE_ENABLED=false (готовые фичи скрыты)"),
    ("Скриншоты отчёта", "живой Playwright с авторизацией (Admin, Cargo Manager, Driver Manager)"),
]
for i, (k, v) in enumerate(checks, start=2):
    put(ws3, i, 1, k, lbl_font, sub_fill); put(ws3, i, 2, v)

out = "C:/Users/hp/Desktop/report/daily-report/10.07.26/report-10.07.26.xlsx"
wb.save(out)
print("saved", out)
