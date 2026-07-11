# Builds report-09.07.26.xlsx — PM summary of 09.07.2026 Sarbon Frontend work (4 sheets).
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = "1639C8"; SOFT = "EEF1FF"; INK = "16181D"; GREEN = "0F9D58"; MUTED = "5B6477"
hdr_fill = PatternFill("solid", fgColor=BLUE)
sub_fill = PatternFill("solid", fgColor=SOFT)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
lbl_font = Font(bold=True, color=INK)
wrap = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="D9DEE8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = wrap; cell.border = border

def put(ws, r, c, v, font=None, fill=None, al=wrap):
    cell = ws.cell(row=r, column=c, value=v)
    if font: cell.font = font
    if fill: cell.fill = fill
    cell.alignment = al; cell.border = border
    return cell

# ---------- Sheet 1: Summary ----------
ws = wb.active; ws.title = "Сводка"
ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 78
put(ws, 1, 1, "Отчёт о работе · 09.07.2026 · Sarbon Frontend", hdr_font, hdr_fill)
put(ws, 1, 2, "Фокус: полная регрессия приложения (8 багов закрыто) · калькуляторы веса/объёма в форме груза · тип звонка в чате · админ-тренд на столбцы", hdr_font, hdr_fill)
rows = [
    ("Дата", "09.07.2026"),
    ("Проект", "sarbon-frontend · React 19 · Vite · Ant Design 6"),
    ("Ветка", "staging"),
    ("Коммиты за день", "4 · 4a207bbf, c79fff0d, 470f44d7, 50851daa"),
    ("Главное", "Трёх-линзовая регрессия всего приложения на реальном staging"),
    ("Багов закрыто", "8 (F-1…F-8), все провалидированы воротами"),
    ("Находок задокументировано", "~35 (готовые фиксы, нужно решение продукта/бэкенда)"),
    ("Ложных срабатываний отсеяно", "5 (адверсариальная проверка против кода)"),
    ("Роли покрыты регрессией", "4 · public, cargo-manager, driver-manager, admin"),
    ("Скриншотов QA-прогона", "526 (54 + 133 + 76 + 263)"),
    ("Новые фичи формы груза", "перерасчёт веса из упаковки · объём из габаритов · подсказки «?» (FieldHelpTooltip)"),
    ("Чат", "распознавание типа звонка (аудио/видео/экран, направление, пропущен)"),
    ("Админ", "тренд-график: линия → столбцы (AdminLine → AdminTrendBars)"),
    ("Профиль", "тёмная тема (dark:-классы карточки)"),
    ("Тесты (pnpm test)", "929/929 (94 файла)"),
    ("Ошибок tsc -b", "0"),
    ("Предупреждений ESLint", "0 (--max-warnings 0)"),
    ("JS-исключений в рантайме", "0 во всех 4 ролевых прогонах"),
    ("Скриншоты-доказательства", "живой Playwright с авторизацией (Cargo Manager +998994878460, Admin)"),
]
for i, (k, v) in enumerate(rows, start=2):
    put(ws, i, 1, k, lbl_font, sub_fill); put(ws, i, 2, v)

# ---------- Sheet 2: Changes ----------
ws2 = wb.create_sheet("Изменения")
widths = [4, 14, 30, 46, 50, 34, 24]
for i, w in enumerate(widths, start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
headers = ["№", "Тип", "Область", "Что было / причина", "Что сделано", "Файлы", "Статус"]
for c, h in enumerate(headers, start=1): put(ws2, 1, c, h)
style_header(ws2, len(headers))
changes = [
    ("QA", "Регрессия всего приложения (3 линзы × 4 роли)",
     "Перед пушем staging→main нужна уверенность по всему приложению, не только по свежим фичам.",
     "Статический fan-out по 9 подсистемам (41 находка) + Playwright-харнесс (526 скриншотов) + визуальный разбор; 8 фиксов, ~35 находок задокументировано, 5 ложных срабатываний отсеяно.",
     "qa/* (конфиги, находки, qa-report-2026-07-09.md), qa_browser.mjs", "✅ 8 фиксов закоммичено"),
    ("feature", "Груз — перерасчёт веса из упаковки",
     "Расхождение «кол-во × вес одной ед.» с полем «Вес» ничем не подсвечивалось.",
     "Из packaging_amount × packaging_unit_weight выводится итог кг/т; при расхождении — жёлтое «не совпадает» + «Применить» (не блокирует сохранение).",
     "utils/packagingWeight.ts (new), Step1CargoInfo.tsx, packagingWeight.test.ts", "✅ Закоммичено"),
    ("feature", "Груз — оценка объёма по габаритам",
     "Объём вводился вручную, хотя часто известны Д×Ш×В.",
     "Из длины×ширины×высоты (м) считается объём габаритного параллелепипеда → «Расчётный объём» + «Применить»; неавторитетно, без автоперезаписи.",
     "utils/dimensionsVolume.ts (new), Step1CargoInfo.tsx, dimensionsVolume.test.ts", "✅ Закоммичено"),
    ("feature", "Форма — подсказки «?» у полей",
     "Не хватало пояснений «что делать, если значение не нашли».",
     "Переиспользуемый FieldHelpTooltip: фокусируемая «?» рядом с ярлыком, тултип по hover/focus/click; контраст выше WCAG 1.4.11.",
     "components/ui/FieldHelpTooltip.tsx (new), fieldHelpTooltip.test.tsx", "✅ Закоммичено"),
    ("feature", "Чат — тип звонка (аудио/видео/экран)",
     "Событие звонка не различало тип — всё рендерилось одинаково.",
     "normalizeChatCallType + разбор тела события; авторитет — payload.call_type, текст — фоллбэк. Пузырь показывает тип + направление + пропущен/завершён.",
     "chat/utils/chatCallEvents.ts, MessageBubble.tsx, ConversationItem.tsx", "✅ Закоммичено"),
    ("redesign", "Админ — тренд-график линия → столбцы",
     "Тренд «Новые водители/менеджеры» рисовался линией — на разрежённых днях читался хуже.",
     "AdminLine → AdminTrendBars: recharts LineChart → BarChart, подписи значений (LabelList), сетка/ось Y; День/Неделя/Месяц как раньше.",
     "charts/AdminTrendBars.tsx (rename), AdminTrendSection.tsx", "✅ Закоммичено"),
    ("polish", "Профиль — тёмная тема",
     "Карточка профиля/«Личные данные» не адаптировались под тёмную тему.",
     "Добавлены dark:-классы для поверхностей, границ и текста карточки профиля и сайдбар-карточки.",
     "profile/ProfilePage.tsx, profile/components/ProfileSidebarCard.tsx", "✅ Закоммичено"),
]
for i, row in enumerate(changes, start=2):
    put(ws2, i, 1, i - 1)
    for c, v in enumerate(row, start=2):
        put(ws2, i, c, v)

# ---------- Sheet 3: Bugs ----------
ws3 = wb.create_sheet("Баги")
widths = [8, 10, 80]
for i, w in enumerate(widths, start=1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
for c, h in enumerate(["ID", "Сев.", "Фикс"], start=1): put(ws3, 1, c, h)
style_header(ws3, 3)
bugs = [
    ("F-1", "High", "Кнопка повторной отправки OTP не показывала/не соблюдала кулдаун — countdownSeconds проброшен в <OtpField> (OtpPage.tsx)."),
    ("F-2", "High", "Повтор OTP при 429 не запускал кулдаун — добавлен startCooldown(retryAfterSeconds) в onError."),
    ("F-3", "High", "Accept в табличном виде офферов терял валюту → встречное предложение уходило в USD — прокинут offer.currency (offersTableColumns.tsx)."),
    ("F-4", "High", "AI-гейт is_cargo: z.coerce.boolean() превращал 'false'/'0' в true — добавлен coerceBoolean (cargoAiParseTypes.ts) + тесты."),
    ("F-5", "High", "Повторный тап Accept звонка утекал живой mic/camera-поток + RTCPeerConnection — acceptInFlightRef-гард (GlobalAudioCallModal.tsx)."),
    ("F-6", "Med", "Дубликаты React-ключей на дашборде — композитный key с индексом (BreakdownPanel.tsx)."),
    ("F-7", "High", "gpsTrackingPage.mapIframeTitle отсутствовал в uz+oz (базовая+фоллбэк локаль) — ключ добавлен."),
    ("F-8", "Low", "Невалидные SVG-пропы fill-rule/clip-rule/clip-path → camelCase (Icons.tsx)."),
]
for i, row in enumerate(bugs, start=2):
    for c, v in enumerate(row, start=1):
        put(ws3, i, c, v)

# ---------- Sheet 4: Checks ----------
ws4 = wb.create_sheet("Проверки")
ws4.column_dimensions['A'].width = 40; ws4.column_dimensions['B'].width = 62
for c, h in enumerate(["Проверка", "Результат"], start=1): put(ws4, 1, c, h)
style_header(ws4, 2)
checks = [
    ("tsc -b (типы, CI-ворота)", "0 ошибок"),
    ("eslint --max-warnings 0", "0 предупреждений (затронутые файлы)"),
    ("pnpm test", "929/929 (94 файла)"),
    ("JS-исключения в рантайме", "0 во всех 4 ролевых прогонах"),
    ("Public / marketing", "9 страниц · 54 скриншота · PASS (визуально чисто)"),
    ("Dispatcher — Cargo Manager", "13 страниц · 133 скриншота · PASS + находки"),
    ("Dispatcher — Driver Manager", "11 страниц · 76 скриншотов · PASS"),
    ("Admin console", "21 страница · 263 скриншота · PASS"),
    ("Адверсариальная проверка находок", "5 ложных срабатываний отсеяно до фикса"),
    ("Скриншоты отчёта", "живой Playwright с авторизацией (staging)"),
]
for i, (k, v) in enumerate(checks, start=2):
    put(ws4, i, 1, k, lbl_font, sub_fill); put(ws4, i, 2, v)

out = "C:/Users/hp/Desktop/report/daily-report/09.07.26/report-09.07.26.xlsx"
wb.save(out)
print("saved", out)
