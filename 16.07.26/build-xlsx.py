# Builds report-16.07.26.xlsx — summary of the day's work: company workspace (/company) +
# admin moderation-log timestamps + admin Drivers/Dispatchers Account-status fix.
# Run: python build-xlsx.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'report-16.07.26.xlsx')

BLUE = '1639C8'; DARK = '0F1D4F'; GREEN = '00B87A'; OKG = '0F9D58'
LINE = 'E6E9F0'; MUTEDBG = 'FAFBFE'
thin = Side(style='thin', color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)
top = Alignment(vertical='top')

wb = Workbook()

# ── Sheet 1: Changes ────────────────────────────────────────────────────────────────
ws = wb.active
ws.title = 'Changes'
ws.sheet_view.showGridLines = False

title = ws.cell(row=1, column=1, value='Sarbon — Company workspace + Admin moderation · 16.07.2026')
title.font = Font(bold=True, size=15, color=DARK)
ws.cell(row=2, column=1,
        value='New /company workspace (17 sections, invitations/roles/fleet/finance, invite-accept + group chat) + admin moderation-log timestamps + Drivers/Dispatchers effective Account status. 145 files / 3 commits.'
        ).font = Font(size=10, color='5B6477', italic=True)
ws.merge_cells('A2:F2')

headers = ['#', 'Type', 'Area', 'Change', 'Endpoint / detail', 'Screenshot']
hr = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=BLUE)
    cell.alignment = Alignment(vertical='center', horizontal='left')
    cell.border = border

rows = [
    ['1', 'FEATURE', 'Company / Workspace',
     '/company/:id turned from a Profile|Cargo toggle into a 17-section grouped sidebar; each section = lazy chunk + pure companyX.ts + unit test, gated by me/permissions',
     'GET /v1/companies/:id, /me/permissions', 'company-overview'],
    ['2', 'FEATURE', 'Company / Team',
     'Invitations (create/list/resend/revoke), employees (role change, per-key permission grant/deny/inherit, remove), roles (retroactive permission overlays)',
     '/companies/:id/{invitations,users,roles}', 'company-employees, -invitations, -roles'],
    ['3', 'FEATURE', 'Company / Fleet',
     'Power + trailer plates CRUD with maintenance start/complete, assignments, maintenance journal, drivers roster (forbidden for SHIPPER)',
     '/companies/:id/fleet/*, /drivers', 'company-vehicles'],
    ['4', 'FEATURE', 'Company / Operations',
     'Company cargo (all statuses), trips (filters + curator-manager assign), finance summary by currency (no FX — profit within one currency)',
     '/companies/:id/{cargo,trips,finance/summary}', 'company-trips, -finance, -cargo'],
    ['5', 'FEATURE', 'Company / Invite + Chat',
     'Invitee side: /accept-invite?token= root page + one-shot breadcrumb through login; chat section launches the company group into the main /chat',
     '/v1/invitations/accept, /companies/:id/group', '—'],
    ['6', 'FIX', 'Company / Profile + antd',
     '"My companies" section (list + switch active company); AntdProvider maps uz -> uz_UZ (was en_US, leaking English antd chrome)',
     'CompanyProfileCompanies, AntdProvider', '—'],
    ['7', 'FEATURE', 'Admin / Moderation log',
     'Separate requested_at + decided_at columns in the moderation log (was a single created_at "Date")',
     'GET /v1/admin/moderation/log', '—'],
    ['8', 'FIX', 'Admin / Drivers',
     'Account column shows effective_status (null -> Active, no more "No data"); actions slimmed to View/Edit/Delete; Account/KYC/Moderation editable in the Edit drawer; activate no longer offered on an already-active row',
     'GET /v1/admin/drivers, /kyc/*, /moderation/*', '01, 03, 04'],
    ['9', 'FIX', 'Admin / Dispatchers',
     'Same pattern keyed on moderation_status (dispatcher account_status is hard-coded active); isBlocked read from moderation_status',
     '/v1/admin/dispatchers/moderation/*', '06, 08, 09, 10'],
]
type_fill = {'FIX': OKG, 'FEATURE': GREEN, 'UI': '7C3AED'}
r = hr + 1
for row in rows:
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = wrap if c in (4, 5, 6) else top
        cell.font = Font(size=10)
        if c == 2:
            cell.font = Font(size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=type_fill.get(val, '64748B'))
            cell.alignment = Alignment(vertical='center', horizontal='center')
    r += 1

widths = [5, 10, 22, 52, 40, 22]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for rr in range(hr + 1, r):
    ws.row_dimensions[rr].height = 54

# ── Sheet 2: Verification ─────────────────────────────────────────────────────────────
ws2 = wb.create_sheet('Verification')
ws2.sheet_view.showGridLines = False
ws2.cell(row=1, column=1, value='Verification gates').font = Font(bold=True, size=14, color=DARK)
gates = [
    ['Gate', 'Result'],
    ['pnpm build (tsc -b + vite)', 'exit 0'],
    ['eslint --max-warnings 0 (changed files)', 'clean'],
    ['vitest run', '1754 / 1754 (165 files)'],
    ['Company workspace (qa/verify-company-workspace.mjs)', '17/17 sections, 0 raw i18n keys'],
    ['Accept-invite (qa/verify-accept-invite.mjs)', '7/7 states'],
    ['Locale parity', '15/15 (company.* en/uz/ru + placeholders)'],
    ['Live run (staging, admin login)', '10 screenshots, behaviour confirmed'],
    ['Files changed', '145 (3 commits)'],
]
gr = 3
for i, (a, b) in enumerate(gates):
    ca = ws2.cell(row=gr + i, column=1, value=a)
    cb = ws2.cell(row=gr + i, column=2, value=b)
    for cell in (ca, cb):
        cell.border = border
        cell.alignment = top
    if i == 0:
        for cell in (ca, cb):
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=BLUE)
    else:
        ca.font = Font(size=10)
        cb.font = Font(size=10, bold=True, color=OKG)
ws2.column_dimensions['A'].width = 48
ws2.column_dimensions['B'].width = 40

wb.save(OUT)
print('wrote', OUT)
