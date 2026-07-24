# Builds report-24.07.26.xlsx — the day's four tickets: GPS driver modal (phone + tractor/trailer
# breakdown), admin Trips joined driver/tractor/last-online/dispatcher columns, cargo creation hidden
# for CARRIER companies, and a share button on the dispatcher My-Cargos table rows. Plus the merge of
# origin/staging into the feature branch so the PR is conflict-free.
# Run: python build-xlsx.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'report-24.07.26.xlsx')

BLUE = '1639C8'; DARK = '0F1D4F'; OKG = '0F9D58'
LINE = 'E6E9F0'
thin = Side(style='thin', color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)
top = Alignment(vertical='top')

wb = Workbook()

# -- Sheet 1: Changes --------------------------------------------------------------------
ws = wb.active
ws.title = 'Changes'
ws.sheet_view.showGridLines = False

ws.cell(row=1, column=1,
        value='Sarbon - GPS driver modal (phone + tractor/trailer breakdown) / admin Trips joined driver+dispatcher columns / cargo creation hidden for CARRIER / share button on My-Cargos rows - 24.07.2026'
        ).font = Font(bold=True, size=14, color=DARK)
ws.cell(row=2, column=1,
        value='Four ticket tasks. GPS driver detail modal now shows the phone and a full tractor/trailer breakdown (fields render only when present). Admin Trips gains Driver/Tractor/Last-online/Dispatcher columns joined by id from the drivers and dispatchers lists (the Trip payload carries only flat ids). The company workspace hides the whole Cargo tab for CARRIER companies (forbiddenForTypes) plus a create-button gate. The dispatcher My-Cargos table rows get the same share menu as All-Cargos. Result merged into a feature branch with origin/staging pulled in, so the PR is conflict-free. Build 0 errors, 2379/2379 tests (203 files), 5 bug.md entries.'
        ).font = Font(size=10, color='5B6477', italic=True)
ws.merge_cells('A2:F2')

headers = ['#', 'Type', 'Area', 'Change', 'Endpoint / detail', 'Screenshot']
hr = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=BLUE)
    cell.alignment = Alignment(vertical='center', horizontal='left')
    cell.border = border

rows = [
    ['1', 'FIX + FEATURE', 'Dispatcher / GPS driver modal',
     'Clicking a driver as a Driver Manager opened a near-empty modal even though GET /v1/dispatchers/drivers/:id returned the data. The phone was never rendered anywhere, and a driver with no route/vehicle left the middle blank. Added a phone card (tel: link, shown only when the phone is available for the role) and a structured Vehicle card with separate Tractor and Trailer blocks (type / model / plate / tech-passport / owner), each row rendered only when present. Also hardened the roster<->detail merge to overlay only non-null fields, so the sparse /drivers/:id no longer wipes roster-known vehicle/rating data.',
     'GET /v1/dispatchers/drivers/:id', '06-gps-driver-modal-phone-vehicle'],

    ['2', 'FEATURE', 'Admin / Trips',
     'The Trip payload (verified against base.yaml#/Trip) carries only flat ids (driver_id, driver_manager_id, ...) plus a vehicle_snapshot - never an embedded driver or dispatcher. Rather than a backend change, joined the admin drivers and dispatchers lists (loaded once, Excel-style) by id and added four columns: Driver (name+phone), Tractor number (vehicle_snapshot.power_plate_number with a driver-plate fallback), Last online (driver.last_online_at), Dispatcher (name+phone via driver_manager_id, falling back to the driver freelancer_id; "no data" when neither resolves). Participants are resolved once in a memoized index so sorting does not re-run the base64 snapshot decode.',
     'GET /v1/admin/trips + /v1/admin/drivers + /v1/admin/dispatchers', '01-admin-trips-driver-dispatcher-columns'],

    ['3', 'FIX', 'Company workspace / Cargo',
     'The company workspace showed the Cargo tab and "Add cargo" to every company. A CARRIER offers transport, not freight, and should not post cargo (the backend rejects the create with company_type_forbidden). The create button was gated only on cargo.create, which fails open for an Owner - so a CARRIER Owner saw a button that only ever 403s. Hid the WHOLE Cargo tab for CARRIER via forbiddenForTypes:[CARRIER] in the workspace registry (it disappears from the nav; a deep-linked ?tab=cargo falls back to the overview), and kept a finer create-button gate (companyTypeAllowsCargoCreation) as defense-in-depth. SHIPPER and BROKER keep the tab.',
     'companyWorkspaceSections registry + CargoSection', '02-company-carrier-no-cargo-tab, 03-company-shipper-cargo-tab'],

    ['4', 'FIX', 'Dispatcher / My cargos',
     'A cargo manager had no way to share their own posting from "My cargos" - the share action existed only in "All cargos". Added the same CargoShareMenu (same ShareCargoToChatModal, same allCargosPage.share.* copy, same endpoints) to the My-Cargos table actions column, rendered on every row; the wrapper stops click propagation so it does not open the row drawer. Scope is the table per the ticket; the card view is a single button wrapper (a nested share button would be invalid button-in-button HTML), left for a future card restructure.',
     'src/pages/cargos/my-cargos/utils/myCargosTableColumns.tsx', '04-my-cargos-share-button, 05-my-cargos-share-menu'],

    ['5', 'GIT', 'Release / merge',
     'While the work was in progress another developer pushed 10 commits to origin/staging (company operations redesign + trips). A plain pull was unsafe (dirty tree + overlap on CargoSection.tsx and the locale files), so the day was pushed to a feature branch and origin/staging was merged INTO it. The only content conflict was the CargoSection signature line (their added i18n vs our added company destructure), resolved by hand; bug.md and the 15 locale files auto-merged with no loss. The branch is now strictly ahead of staging (0 behind / 3 ahead) so the PR merges cleanly.',
     'feat/gps-modal-admin-trips-carrier-share-2026-07-24', '(git)'],
]
r = hr + 1
for i, row in enumerate(rows):
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r + i, column=c, value=val)
        cell.border = border
        cell.alignment = wrap
        cell.font = Font(size=10, bold=(c == 1))
        if c == 2:
            cell.font = Font(size=10, bold=True, color=OKG if 'FEATURE' in str(val) else BLUE)
widths = [4, 14, 26, 96, 46, 34]
for c, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + c)].width = w
for rr in range(r, r + len(rows)):
    ws.row_dimensions[rr].height = 118

# -- Sheet 2: Verification ---------------------------------------------------------------
ws2 = wb.create_sheet('Verification')
ws2.sheet_view.showGridLines = False
ws2.cell(row=1, column=1, value='Verification & gates - 24.07.2026').font = Font(bold=True, size=14, color=DARK)
gates = [
    ['Gate', 'Result'],
    ['Build (tsc -b + vite)', '0 errors'],
    ['ESLint --max-warnings 0 (changed files)', 'clean'],
    ['Unit tests (pnpm test, after merge)', '2379 / 2379 passed (203 files)'],
    ['New pure modules + tests', 'adminTripParticipants (19) + companyCargoForm (+3) + companyWorkspaceSections (+2)'],
    ['Locales', '15/15 parse with the new gpsTrackingPage.* and admin.pages.trips.table.* keys'],
    ['GPS driver modal', 'Playwright: phone card + tractor/trailer breakdown render on a seeded DRIVER_MANAGER session (Yandex map behind)'],
    ['Admin Trips columns', 'Playwright: driver/tractor/last-online/dispatcher columns populate; "no data" for the driver with no dispatcher'],
    ['CARRIER vs SHIPPER', 'Playwright: Cargo tab absent for CARRIER, present for SHIPPER'],
    ['My Cargos share', 'Playwright: share icon on every table row + the share menu opens'],
    ['Live authenticated staging run', 'NOT run - the four surfaces are auth-gated; captured on seeded sessions + stubbed API instead'],
    ['Merge / PR', 'origin/staging merged into the feature branch; branch 0 behind / 3 ahead -> PR conflict-free'],
]
gr = 3
for i, (a, b) in enumerate(gates):
    ca = ws2.cell(row=gr + i, column=1, value=a)
    cb = ws2.cell(row=gr + i, column=2, value=b)
    for cell in (ca, cb):
        cell.border = border
        cell.alignment = wrap if i > 0 else top
    if i == 0:
        for cell in (ca, cb):
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=BLUE)
    else:
        ca.font = Font(size=10)
        bad = b.startswith('NOT ')
        cb.font = Font(size=10, bold=True, color='E8590C' if bad else OKG)
ws2.column_dimensions['A'].width = 42
ws2.column_dimensions['B'].width = 96
for rr in range(gr, gr + len(gates)):
    ws2.row_dimensions[rr].height = 34

# -- Sheet 3: Screenshots ----------------------------------------------------------------
ws3 = wb.create_sheet('Screenshots')
ws3.sheet_view.showGridLines = False
ws3.cell(row=1, column=1, value='Screenshots - all captured with Playwright against the running app (feature branch)').font = Font(bold=True, size=14, color=DARK)
ws3.cell(row=2, column=1,
         value='Every surface is behind authentication, so each shot runs on a seeded session (dispatcher / admin / company token in localStorage) with the relevant staging responses stubbed - the same harness family used to verify the changes. No live staging credentials were used.'
         ).font = Font(size=10, color='5B6477', italic=True)
ws3.merge_cells('A2:C2')
shots = [
    ['File', 'Shows', 'Mode'],
    ['06-gps-driver-modal-phone-vehicle.png', 'GPS driver modal: phone card + Tractor/Trailer breakdown (type/model/plate/tech-passport/owner)', 'seeded DRIVER_MANAGER + stubbed'],
    ['01-admin-trips-driver-dispatcher-columns.png', 'Admin Trips: Driver (name+phone), Tractor number, Last online, Dispatcher columns joined by id', 'seeded admin creator + stubbed'],
    ['02-company-carrier-no-cargo-tab.png', 'CARRIER company workspace nav - Cargo tab absent (Fleet group present)', 'seeded company + stubbed'],
    ['03-company-shipper-cargo-tab.png', 'SHIPPER company workspace nav - Cargo tab present (Fleet group absent)', 'seeded company + stubbed'],
    ['04-my-cargos-share-button.png', 'My cargos table (Cargo Manager) - share icon in every row action cell', 'seeded dispatcher + stubbed'],
    ['05-my-cargos-share-menu.png', 'My cargos - the share menu open (Chat Sarbon / copy link / WhatsApp / Telegram / ...)', 'seeded dispatcher + stubbed'],
    ['06b-gps-driver-modal-full.png', 'Full-page GPS view with the driver modal open over the map', 'seeded DRIVER_MANAGER + stubbed'],
]
sr = 4
for i, row in enumerate(shots):
    for c, val in enumerate(row, start=1):
        cell = ws3.cell(row=sr + i, column=c, value=val)
        cell.border = border
        cell.alignment = wrap
        if i == 0:
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=BLUE)
        else:
            cell.font = Font(size=10)
ws3.column_dimensions['A'].width = 46
ws3.column_dimensions['B'].width = 86
ws3.column_dimensions['C'].width = 32
for rr in range(sr, sr + len(shots)):
    ws3.row_dimensions[rr].height = 30

wb.save(OUT)
print('wrote', OUT)
