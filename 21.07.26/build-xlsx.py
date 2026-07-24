# Builds report-21.07.26.xlsx — the day's work: a page-by-page audit and repair pass over the whole
# company surface (workspace, roles, audit log, auth, create, profile), admin company moderation,
# the command-center moderation queues, the dispatcher personal<->company context switcher, and an
# end-of-day self-review that found 12 more defects in the day's own fixes.
# Run: python build-xlsx.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'report-21.07.26.xlsx')

BLUE = '1639C8'; DARK = '0F1D4F'; GREEN = '00B87A'; OKG = '0F9D58'
LINE = 'E6E9F0'
thin = Side(style='thin', color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)
top = Alignment(vertical='top')

wb = Workbook()

# -- Sheet 1: Changes --------------------------------------------------------------------
ws = wb.active
ws.title = 'Changes'
ws.sheet_view.showGridLines = False

title = ws.cell(row=1, column=1,
                value='Sarbon - Company surface page-by-page audit & repair / admin company moderation / command center / context switcher - 21.07.2026')
title.font = Font(bold=True, size=15, color=DARK)
ws.cell(row=2, column=1,
        value='Three commits (630985a1 -> a4a647eb -> 1a829081) plus working-tree work. Almost the whole day was a screen-by-screen pass over the company surface - 16 bug.md entries, each with a root cause and a live re-verification. Plus four new moderation queues in the admin command center, the dispatcher personal<->company context switcher, and an end-of-day self-review that found 12 further defects in the day\'s own fixes. 83 files in src/ and locales (+2651 / -1257).'
        ).font = Font(size=10, color='5B6477', italic=True)
ws.merge_cells('A2:F2')

headers = ['#', 'Type', 'Area', 'Change', 'Endpoint / detail', 'Screenshot']
hr = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=BLUE)
    cell.alignment = Alignment(vertical='center', horizontal='left')
    cell.border = border

rows = [
    ['1', 'FIX', 'Company workspace / Overview',
     'Company name + avatar rendered three times in one page tree (page h1, sidebar header, overview card title); the sidebar identity header was removed and the overview card retitled. The overview showed fleet-shaped data (max_vehicles/max_trailers/max_drivers seat tiles, vehicle/trailer/driver quota tiles, fleet-maintenance alert) for SHIPPER companies that structurally have no fleet - the nav already hid the whole fleet group for them. Added companyTypeHidesFleet + visibleCompanySeatLimitEntries / visibleCompanyQuotaItems keyed on the same type token the nav gate uses',
     'GET /v1/companies/:id, .../quota, .../dashboard',
     '01-company-overview-carrier, 06-company-overview-shipper-no-fleet, 06f-company-overview-shipper-full'],

    ['2', 'FIX', 'Company workspace / Audit log',
     'The audit tab printed raw backend enum codes (create, user_company_role, company_user) and the detail drawer dumped JSON.stringify(old_data/new_data) plus full UUIDs - unreadable to an owner. Added prettifyAuditToken, action/entity/actor label dictionaries and diffCompanyAuditData; the drawer is now a "what changed: before -> after" field list. New i18n subtrees in all 15 locales',
     'GET /v1/companies/:id/audit',
     '04-company-audit-humanized, 05b-company-audit-change-drawer'],

    ['3', 'UI', 'Company workspace / Light theme',
     'Day mode looked empty - near-white surfaces on a near-white page separated only by a faint 1px border, with no elevation system at all (--sarbon-surface #fff vs --sarbon-bg #edeff5 is a tiny step). Added a theme-aware --sarbon-shadow-card token (soft two-layer in light, near-nil in dark) applied to the section cards, the workspace sidebar and the page skeletons',
     'src/index.css', '(visible on every company screenshot)'],

    ['4', 'FIX', 'Company workspace / Roles',
     'Reported by the product owner: setting a permission to Allowed/Denied and pressing Save showed a success toast, then every toggle snapped back to "Not set". FRONTEND behavior with a BACKEND gap underneath - each toggle read only the local pending map and onSuccess did setPending({}), wiping the single source it read; and the role-permission API was verified WRITE-ONLY against the live spec (PUT only, no GET), so the FE could never read back what it saved. Session mitigation via resolveRoleSegmentValue / mergeAppliedRoleOverlay / removeOverlayKey; backend ticket filed for GET .../roles/{role}/permissions (delivered the next morning)',
     'PUT /v1/companies/:id/roles/:role/permissions (no GET at the time)',
     '03c-company-roles-persisted-after-save'],

    ['5', 'FIX', 'Company workspace / Roles',
     'The tab always offered all 7 configurable roles regardless of company type, including the three fleet roles (SeniorDriverManager, DriverManager, FleetManager) for SHIPPER companies with no fleet - configuring them could only ever be rejected as role_not_allowed_for_company_type. Added isRoleAvailableForCompanyType + configurableRolesForCompanyType; an unknown/absent type falls open (shows all) rather than hiding by guess',
     'companyRoles.ts', '02-company-roles-carrier-all-roles, 07-company-roles-shipper-filtered'],

    ['6', 'FIX', 'Company chat / Sidebar footer',
     'On /company/:id/chat the sidebar\'s bottom-left own-account row showed a "?" avatar and the label "You". ChatSidebarPanel is part of the reused dispatcher ChatPage and its footer was the one place reading the identity straight from the dispatcher store, bypassing the surface-aware useChatIdentity - empty on the company surface (and, with a parallel dispatcher session in the browser, it would have leaked THAT identity instead). Added resolveChatSelfDisplay + useChatSelfDisplay; dispatcher /chat unchanged',
     'src/pages/chat/{utils/chatSelfDisplay.ts,hooks/useChatSelfDisplay.ts}', '13-company-chat-self-footer'],

    ['7', 'FIX', 'Company auth / Accept invite (CRITICAL)',
     'An authenticated invitee opening an invite link was trapped in an endless accept-invite -> company/login -> company -> accept-invite cycle: address bar flapping, permanent spinner, no way out but closing the tab. Measured before the fix: 844 accept POSTs and 242 navigations in 8 seconds. classifyAcceptInvitationError falls back to login_required on any marker-less 401 (correct for a guest) and onError detoured unconditionally; with a live session the guard bounces the user back and the page remounts with a fresh attemptedRef. New pure predicate shouldDetourToLogin(kind, isAuthenticated). After: 2 POSTs, 2 navigations, a stable error card',
     'POST /v1/invitations/:id/accept', '(measured, not re-shot - needs a live invite token)'],

    ['8', 'FIX', 'Company session / Cross-tab (CRITICAL)',
     'After switching company in tab A, tab B kept rendering company A while its requests already carried company B\'s JWT - either a 403 not_member_of_company or, worse, plausible data attributed to the wrong company. The cross-tab storage listener fired only on COMPANY_TOKEN, but syncAuth() reads the context and profile too, and the switch writes token first / context second - so the passive tab re-read the old context and ignored the context write, activeCompany never changed, and clearCompanyScopedQueries never ran. Now every key syncAuth() reads is watched; write order deliberately left alone',
     'useCompanyAuthStore syncAuth / POST /v1/auth/switch-company', '(cross-tab, not screenshot-able)'],

    ['9', 'FIX', 'Admin / Companies (CRITICAL)',
     'The row action "Bind owner" (and the post-create hand-off) opened the Assign tab with a blank Company ID - it worked only if that tab had already been visited in the session. AntD mounts a tab pane on first activation, so the imperative assignForm.setFieldValue ran against an unconnected form instance and was dropped. Fixed with forceRender: true on the assign tab; costs no extra request (owner search is gated on 2+ characters)',
     'AdminCompaniesPage.tsx', '23-admin-companies-row-kebab'],

    ['10', 'FEATURE', 'Admin / Command center',
     '"Needs attention" only surfaced cargo + drivers from totals and pulled manager counts from a separate /v1/admin/analytics/moderation call. The backend now returns every queue count in totals, so the redundant query (with its per-row spinner and em-dash degrade) was removed and two rows added: Companies (-> ?tab=company) and Driver KYC (-> ?tab=passport); cargo now links to its own tab too. The bar takes height "100%" inside an lg:items-stretch cell so it fills the now-taller six-row left column instead of a fixed 176px',
     'GET /v1/admin/dashboard?bucket=day', '20-admin-command-center-attention'],

    ['11', 'FIX', 'Company auth / Login, register, OTP, complete',
     'Four screens in a row. Login: a failed login was silent, raw backend codes reached the toast, an empty field produced the wrong error, and the method tabs were unusable by keyboard. Register: resend downgraded the registration flow to a login; hand-built lang path. OTP: silent dead-end on a malformed 200, and network errors burned attempts toward the 5-strike lockout (isTransientApiError now short-circuits). Complete: "already registered" left the user on a form that could only 409 forever - now navigates to the company login with the identity prefilled via router state (phone stays out of the URL and history)',
     'POST /v1/company-users/auth/{phone,otp/verify,login/password}, .../registration/complete',
     '11b-company-login-localized-error, 12b-company-otp-guard-email'],

    ['12', 'FIX', 'Auth / OTP resend cooldown',
     'The countdown was ONE global timer for every identifier and every auth surface: a 429 on one number disabled the button for any other number and for all five dispatcher screens at once, and a pristine company auth page rendered the dispatcher\'s countdown. Rewritten as an identifier-scoped cooldown (no argument -> global, an identifier -> its own scope, null -> an isolated __pending__ scope)',
     'src/pages/auth/hooks/useOtpCooldown.ts', '11a-company-login'],

    ['13', 'FIX', 'Company / Create',
     'A taken INN surfaced as the literal token already_exists: the endpoint answers a duplicate with a bare 409 and no data.fields, so the field-error mapping never caught it and the raw description reached the toast, untranslated and unattached to the input. New isCompanyInnTakenError pins a localized message to the INN field. Also: an unreadable 200 stranded the user on the form with the company ALREADY created (a resubmit could only 409 on the same INN) - now warns/announces, invalidates the list and lands on it; and the form stayed interactive during the follow-up extras PATCH, so it was resubmittable mid-navigation',
     'POST /v1/companies, PATCH /v1/companies/:id', '10c-company-create-inn-taken-field'],

    ['14', 'FIX', 'Company / Profile',
     'company.profile.companies.rating.{overall,shipper,carrier} existed in NO locale (the only key family on the whole company surface that actually leaked raw), so any company with a numeric rating printed the key itself - added to all 15. The active-company card could lock into an unrecoverable 403: resolveActiveCompany ends in "first company", a display heuristic and not a JWT scope guarantee, so on a plain login the detail fetch 403\'d forever behind a Retry that could never succeed - new isCompanyDetailFetchable gates the fetch on the server\'s own is_current, and a 403 that does happen renders the switch hint instead. Email-registered accounts saw "-" as their login identity',
     'GET /v1/companies/:id, PATCH /v1/company-users/profile',
     '09-company-profile-ratings-email, 09b-company-profile-companies-ratings'],

    ['15', 'FIX', 'Company / Full read-only audit',
     'HIGH: CargoSection stale closure - offerColumns useMemo omitted offersCargo from its deps behind an eslint-disable, and the offers drawer opens by setting exactly that state, so the mount-time closure was reused and clicking Accept as the first action silently no-op\'d (no call, no spinner, no error). HIGH: silent no-op on a malformed 200 in OTP + login onSuccess. MED: create double-submit -> duplicate company; unguarded OTP screen; transient errors misclassified as a wrong code; roles revert race. Plus LOW cleanups and the AntD 6 Alert message->title deprecation across 11 company files',
     '18 changed files', '(logic - covered by the company suites)'],

    ['16', 'FIX', 'Admin / Company moderation',
     'Restore after a soft delete was unreachable: the list endpoint cannot show deleted companies (verified against the live spec), so once the row refetched away POST .../restore had no UI entry point at all - the one moment it IS reachable is right after the delete, so the success now shows a 10s notification with a Restore button. Stale Excel filters blanked the table after a Segmented status switch (both distinctColumnFilter columns were uncontrolled while the Segmented swaps the whole server dataset) - wired the documented pruneFilteredValue pattern. The moderation hub had no tariff access, the only mechanism for a company\'s max_* limits - AssignTariffModal now mounted there too. created_at rendered 5 hours behind (formatDateTimeIsoClock ignores the offset) -> formatCargoRouteDateTime. Create form: status was free text -> Select, inn had no format rule -> 9/14-digit validation, max_trailers added (7 of 8 limits were present); owner search debounced 300ms; "set default" tariff now behind a confirm',
     'GET /v1/admin/companies, POST .../restore, POST .../tariff',
     '21-admin-companies-queue, 22-admin-companies-column-filter, 24-admin-companies-create-validation, 25-admin-companies-tariffs, 26/27-admin-moderation-company-tab'],

    ['17', 'FEATURE', 'Dispatcher / Context switcher',
     'A dispatcher hired by a company can now switch working context (personal <-> company) from a header dropdown, with a "My company" group in the sidebar. POST /v1/dispatchers/switch-context issues a NEW token pair - it changes the SESSION at JWT level, not the screen - so activeContext is persisted in useAuthStore (DISPATCHER_CONTEXT key), synced cross-tab, and useDispatcherContextSync drops the cache on change, mirroring the owner-side company switch. A dispatcher with no memberships sees no control at all. Same commit: invite-flow name prefill (companyRegisterPrefill) and intent=invite so an invitee lands on the invitation inbox instead of the create-a-company step',
     'GET /v1/dispatchers/companies, POST /v1/dispatchers/switch-context',
     '30-dispatcher-context-switcher, 31-dispatcher-sidebar-my-company'],

    ['18', 'REFACTOR', 'Repo cleanup',
     'Removed dead code and scratch artifacts: the OTP-channel generator (imported by nothing but its own test), three unused command-center charts (Lollipop, RadialBar, Leaderboard + its mapping), an unused admin notifications button and analytics metric card, and ~23k lines of temporary spec dumps, xlsx exports, a dist.zip and stray stack dumps from the repo root',
     '1a829081', '(no UI)'],

    ['19', 'FIX', 'Self-review of the day\'s own fixes',
     'An 8-angle review over the uncommitted company diff; 12 findings survived, all fixed. Worst: isMachineErrorCode suppressed any lowercase-ASCII description - but uz is the platform default language and IS lowercase Latin, so a legitimate one-word localized message ("bloklangan", "topilmadi") was indistinguishable from a code and got replaced by the generic fallback. Others: the restore undo could never work (TanStack Query v5 skips mutate-level callbacks once the row unmounts); the countdown collapsed after 1s with storage blocked; typing wiped the repeated-failure marker; a 429 or expired session counted as a wrong OTP code; a 409 with field errors was relabelled "INN taken"; an unreadable 200 claimed success; any 403 was reported as "wrong company scope"; an explicit is_current:false read as "scope unknown" and opened the gate onto the very 403 the guard prevents',
     'useOtpCooldown, companyAuthError, companyCreate, companyActive, httpStatus (new)',
     '(logic - covered by the unit suites)'],

    ['20', 'OPEN', 'Company auth / OTP guard (phone channel)',
     'FOUND WHILE SHOOTING THIS REPORT, NOT FIXED: CompanyOtpPage computes phone = ensurePlusPrefix(searchParams.get("phone") || ""), which yields "+" when the param is absent. "+" is truthy, so the !identifier guard never fires and a direct visit to /company/otp with no phone still renders a live code screen addressed to a lone "+". The email channel guards correctly. Fix: test the raw param, or treat a bare "+" as empty',
     'src/pages/company/auth/otp/CompanyOtpPage.tsx', '12-company-otp-guarded (shows the hole)'],
]
type_fill = {'FIX': OKG, 'FEATURE': GREEN, 'UI': '7C3AED', 'REFACTOR': '64748B', 'OPEN': 'E8590C'}
r = hr + 1
for row in rows:
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = wrap if c in (4, 5, 6) else top
        cell.font = Font(size=10)
        if c == 2:
            cell.font = Font(size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=type_fill.get(val, '64748B'))
            cell.alignment = Alignment(vertical='center', horizontal='center')
    r += 1

widths = [5, 11, 26, 62, 38, 34]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for rr in range(hr + 1, r):
    ws.row_dimensions[rr].height = 96

# -- Sheet 2: Verification ---------------------------------------------------------------
ws2 = wb.create_sheet('Verification')
ws2.sheet_view.showGridLines = False
ws2.cell(row=1, column=1, value='Verification gates - 21.07.2026').font = Font(bold=True, size=14, color=DARK)
gates = [
    ['Gate', 'Result'],
    ['tsc -b (full project type-check)', 'exit 0'],
    ['pnpm build', 'exit 0'],
    ['vitest (end of day)', '187 files / 2075 tests'],
    ['eslint --max-warnings 0', 'clean on every changed file'],
    ['Locale parity', '15/15 (profile ratings, audit labels, tariff set-default confirm, create unclearResult, command-center queues)'],
    ['qa/verify-company-{workspace,surface,create}.mjs', 'exit 0; create -> "ALL CHECKS PASS"'],
    ['Live re-verification (Playwright)', 'cooldown scopes, delete->restore undo, queue filter pruning, 403 gate, role permissions, company-type gating'],
    ['Accept-invite loop with a live session', '844 POSTs / 242 navigations -> 2 POSTs / 2 navigations'],
    ['Read-back of saved role permissions', 'BLOCKED on 21.07 (no GET) - ticket filed; backend shipped the GET on the morning of 22.07'],
    ['OTP guard on the phone channel', 'OPEN - ensurePlusPrefix("") returns "+", so the guard never fires (found while shooting this report)'],
    ['Files touched', '83 in src/ and locales (+2651 / -1257), 3 commits + working tree'],
    ['bug.md entries', '16 for the day'],
]
gr = 3
for i, (a, b) in enumerate(gates):
    ca = ws2.cell(row=gr + i, column=1, value=a)
    cb = ws2.cell(row=gr + i, column=2, value=b)
    for cell in (ca, cb):
        cell.border = border
        cell.alignment = wrap if i > 0 else top
    if i == 0:
        for cell in (ca, cb):
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=BLUE)
    else:
        ca.font = Font(size=10)
        bad = ('BLOCKED' in b) or ('OPEN' in b)
        cb.font = Font(size=10, bold=True, color='E8590C' if bad else OKG)
ws2.column_dimensions['A'].width = 46
ws2.column_dimensions['B'].width = 78
for rr in range(gr, gr + len(gates)):
    ws2.row_dimensions[rr].height = 34

# -- Sheet 3: Screenshots -----------------------------------------------------------------
ws3 = wb.create_sheet('Screenshots')
ws3.sheet_view.showGridLines = False
ws3.cell(row=1, column=1, value='Screenshots - all captured with Playwright against the running app').font = Font(bold=True, size=14, color=DARK)
ws3.cell(row=2, column=1,
         value='Admin screens run against live staging under a real creator-admin session. Company-owner screens run on a seeded session with stubbed staging responses - the same harness the fixes were verified with during the day, because the owner surface has no shareable test account. The delete -> Restore undo notification was deliberately NOT re-shot: it would soft-delete a real staging company.'
         ).font = Font(size=10, color='5B6477', italic=True)
ws3.merge_cells('A2:C2')
shots = [
    ['File', 'Shows', 'Mode'],
    ['01-company-overview-carrier.png', 'Carrier workspace overview - fleet nav group present, identity rendered once, card elevation', 'seeded + stubbed'],
    ['06-company-overview-shipper-no-fleet.png', 'Shipper workspace overview - fleet nav group absent', 'seeded + stubbed'],
    ['06f-company-overview-shipper-full.png', 'Shipper full page - fleet seat/quota tiles and the maintenance alert filtered out', 'seeded + stubbed'],
    ['02-company-roles-carrier-all-roles.png', 'Roles tab, carrier - all 7 configurable roles', 'seeded + stubbed'],
    ['07-company-roles-shipper-filtered.png', 'Roles tab, shipper - 4 roles, the three fleet roles hidden', 'seeded + stubbed'],
    ['03c-company-roles-persisted-after-save.png', 'A saved permission stays on Allowed/Denied instead of snapping back to "Not set"', 'seeded + stubbed'],
    ['04-company-audit-humanized.png', 'Audit log with humanized action / entity / actor columns', 'seeded + stubbed'],
    ['05b-company-audit-change-drawer.png', '"What changed" drawer - before -> after per field, no raw JSON, no UUIDs', 'seeded + stubbed'],
    ['13-company-chat-self-footer.png', 'Company chat sidebar footer showing the signed-in company user', 'seeded + stubbed'],
    ['11b-company-login-localized-error.png', 'Company login - localized 401 plus both identity and method toggles', 'seeded + stubbed'],
    ['12b-company-otp-guard-email.png', 'OTP screen guard on the email channel - warning + disabled form', 'seeded + stubbed'],
    ['12-company-otp-guarded.png', 'OPEN FINDING: the same guard does not fire on the phone channel', 'seeded + stubbed'],
    ['10c-company-create-inn-taken-field.png', 'Taken INN reported on the INN field, not as a raw already_exists toast', 'seeded + stubbed'],
    ['09-company-profile-ratings-email.png', 'Company profile - email as the login identity (was a dash)', 'seeded + stubbed'],
    ['09b-company-profile-companies-ratings.png', 'My companies - localized rating labels (previously raw i18n keys)', 'seeded + stubbed'],
    ['20-admin-command-center-attention.png', 'Command center "Needs attention" - six queues, bar stretched to column height', 'live staging'],
    ['21-admin-companies-queue.png', 'Admin company moderation queue with local-time "Created" column', 'live staging'],
    ['22-admin-companies-column-filter.png', 'Status column filter offering only values present in the current rows', 'live staging'],
    ['23-admin-companies-row-kebab.png', 'Row actions - Bind owner / Assign tariff seeded with the right company id', 'live staging'],
    ['24-admin-companies-create-validation.png', 'Create form - INN 9/14-digit rule, status Select, max_trailers present', 'live staging'],
    ['25-admin-companies-tariffs.png', 'Creator-only tariffs catalog', 'live staging'],
    ['26-admin-moderation-company-tab.png', 'The same queue mounted in the moderation hub', 'live staging'],
    ['27-admin-moderation-company-kebab-tariff.png', '"Assign tariff" now reachable from the hub row kebab', 'live staging'],
    ['30-dispatcher-context-switcher.png', 'Header personal <-> company context switcher', 'live staging + stubbed memberships'],
    ['31-dispatcher-sidebar-my-company.png', 'Sidebar "My company" group for a hired dispatcher', 'live staging + stubbed memberships'],
]
sr = 4
for i, row in enumerate(shots):
    for c, val in enumerate(row, start=1):
        cell = ws3.cell(row=sr + i, column=c, value=val)
        cell.border = border
        cell.alignment = wrap
        if i == 0:
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=BLUE)
        else:
            cell.font = Font(size=10)
ws3.column_dimensions['A'].width = 46
ws3.column_dimensions['B'].width = 78
ws3.column_dimensions['C'].width = 32
for rr in range(sr, sr + len(shots)):
    ws3.row_dimensions[rr].height = 30

wb.save(OUT)
print('wrote', OUT)
