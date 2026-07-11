# Builds report-07.07.26.xlsx — PM summary of 07.07.2026 Sarbon Frontend work (4 sheets).
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = "1639C8"; SOFT = "EEF1FF"; INK = "16181D"; GREEN = "0F9D58"; MUTED = "5B6477"
hdr_fill = PatternFill("solid", fgColor=BLUE)
sub_fill = PatternFill("solid", fgColor=SOFT)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
lbl_font = Font(bold=True, color=INK)
wrap = Alignment(wrap_text=True, vertical="top")
top = Alignment(vertical="top")
thin = Side(style="thin", color="D9DEE8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = wrap; cell.border = border

def put(ws, r, c, v, font=None, fill=None, al=wrap):
    cell = ws.cell(row=r, column=c, value=v)
    if font: cell.font = font
    if fill: cell.fill = fill
    cell.alignment = al; cell.border = border
    return cell

# ---------- Sheet 1: Summary ----------
ws = wb.active; ws.title = "Сводка"
ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 70
put(ws, 1, 1, "Отчёт о работе · 07.07.2026 · Sarbon Frontend", hdr_font, hdr_fill)
put(ws, 1, 2, "Фокус: AI-создание груза · переработка авторизации · чат-рекордер под Telegram · паспортная модерация · i18n", hdr_font, hdr_fill)
rows = [
    ("Дата", "07.07.2026"),
    ("Проект", "sarbon-frontend · React 19 · Vite · Ant Design 6"),
    ("Ветка", "staging"),
    ("Коммиты за день", "4 · 5e1683ba, d597b2ff, fa72742b, d04332f8 (+ рабочее дерево)"),
    ("Направлений", "6"),
    ("Новая страница", "AI-создание груза (/cargo-ai-create)"),
    ("Дефектов закрыто", "3 (email+пароль, флаг телефона, дубль водителя KYC)"),
    ("Фиксов", "6 (+5 ревью-фиксов AI)"),
    ("Тесты (vitest)", "822/822 на этапе AI-создания + новые сьюты (auth/moderation/cargoAi/chat/offers)"),
    ("Ошибок tsc -b", "0 (затронутые файлы)"),
    ("Предупреждений ESLint", "0 (--max-warnings 0)"),
    ("Локали", "15/15 валидны (+offerType, +auth/AI-ключи)"),
    ("Скриншоты-доказательства", "авторизация — живой Playwright (ru); AI/офферы — репродукции реальных классов"),
    ("Runtime-верификация", "auth ✅ живьём (публичные страницы); AI/чат/админ — на staging за логином"),
]
for i, (k, v) in enumerate(rows, start=2):
    put(ws, i, 1, k, lbl_font, sub_fill); put(ws, i, 2, v)

# ---------- Sheet 2: Changes ----------
ws2 = wb.create_sheet("Изменения")
widths = [4, 16, 30, 46, 46, 30, 30]
for i, w in enumerate(widths, start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
headers = ["№", "Тип", "Область", "Что было / причина", "Что сделано", "Файлы", "Статус"]
for c, h in enumerate(headers, start=1): put(ws2, 1, c, h)
style_header(ws2, len(headers))
changes = [
    ("feature", "Груз — AI-создание груза (новая страница)",
     "Создание груза шло только вручную через 4-шаговый мастер; объявления приходят текстом в мессенджерах.",
     "Страница /cargo-ai-create: вставка текста → DeepSeek парсит → автозаполнение; зелёный/жёлтый/красный по уверенности, ревёрт к AI, повторное распознавание, черновик в localStorage (24ч), предупреждение о дубликате; телефоны/цены редактируются в comment; вывод zod-валидируется; ключ за dev-прокси.",
     "cargo-ai-create/*, cargo-create/{context,hooks,utils}/*, api/ai/*, 5 тестов, vite.config.ts, локали ×15",
     "✅ Закоммичено · build/lint ✅ · vitest 822/822 · DeepSeek 200"),
    ("bug", "Авторизация — «email + пароль»",
     "Выбор Email+Пароль слал {email,password} без телефона → бэкенд «Invalid request body» (пароль только по телефону).",
     "Email привязан к методу «код»: identity = method==='password' ? 'phone' : selected (синхронно, без фликера), email-опция disabled на пароле; onError по variables.data (нет гонки).",
     "LoginPage.tsx, Auth{Identity,OtpChannel,OtpIdentity}*, authEmail/authLoginIdentity/authOtpChannel.ts",
     "✅ Закоммичено · 32 auth-теста ✅"),
    ("bug", "Авторизация — угол флаг-кнопки телефона",
     "Квадратный угол кнопки страны выступал за скруглённую 2px-рамку — рамка «резала» фон кнопки.",
     "В index.css убран min-height:48px !important + добавлен overflow:hidden (клипит только флаг; дропдаун — сиблинг). Одна общая правка для всех PhoneInput.",
     "src/index.css (~L1267)",
     "✅ Закоммичено · Runtime-verified Playwright (light/dark)"),
    ("redesign", "Чат — рекордер под Telegram (голос ⇄ видеокружок)",
     "Две кнопки; голос = инлайн-бар, видео = полноэкранная модалка; не было единой модели и жеста удержания.",
     "Один тумблер голос⇄видео + единая панель записи + плавающее круг-превью; useHoldToRecordGesture (свайп-вверх — блок, влево — отмена); rAF-таймеры; харднинг флипа камеры / обрезки (timeslice 1с) / «не пишет» (onerror).",
     "useVoiceRecorder.ts, useVideoNoteRecorder.ts, useHoldToRecordGesture.ts (new), ChatComposerPanel/ChatVideoNoteCirclePreview/ChatThreadPanel/ChatPage.tsx",
     "✅ Закоммичено · vitest 776/776"),
    ("feature+bug", "Админ — паспортная модерация вкладкой + фасет типа + фикс дубля",
     "Вкладка модерации имела только груз/водители/менеджеры (все — допуск); паспортная (identity) модерация потеряна. Один водитель рендерился сотнями строк.",
     "Отдельная вкладка «Passport moderation» (Водители — /kyc/*; Менеджеры — best-effort PATCH /dispatchers/:id) + фасет «тип модерации» в логе. Фикс дубля: KYC-очередь переведена на paging:'offset' (limit/offset).",
     "PassportModerationPanel.tsx (new), passportModeration.ts (new), AdminModerationPage.tsx, moderationLogUtils.ts, AdminModerationLogPage.tsx, локали ×15",
     "✅ Закоммичено · 13 новых + 23 существующих теста"),
    ("bug", "Офферы — тип «Simple» по-английски на всех языках",
     "offerTypeLabel резолвит offerType.${KEY}, но группа offerType отсутствовала во всех 15 локалях → падение на англ. defaultValue «Simple».",
     "Добавлена группа offerType (SIMPLE — «обычное предложение», COUNTER — зеркалит offersPanel.counterOffer) во все 15 локалей. Кода не трогали. Админ-офферы не затронуты (свой namespace).",
     "public/locales/*.json ×15 (+offerType)",
     "🛠 Рабочее дерево · offersPageHelpers 9/9 ✅"),
]
for i, row in enumerate(changes, start=2):
    put(ws2, i, 1, i - 1, top)
    for c, v in enumerate(row, start=2):
        put(ws2, i, c, v)

# ---------- Sheet 3: Files ----------
ws3 = wb.create_sheet("Файлы")
for i, w in enumerate([54, 20, 52, 22], start=1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
for c, h in enumerate(["Файл", "Тип изменения", "Что сделано", "Коммит"], start=1): put(ws3, 1, c, h)
style_header(ws3, 4)
files = [
    ("src/pages/cargos/cargo-ai-create/CargoAiCreatePage.tsx", "UI (new)", "Страница AI-создания: paste-box, AI-обзор, автозаполнение", "5e1683ba"),
    ("src/api/ai/deepseekCargoParser.ts", "Логика (new)", "parseCargoAnnouncement() + zod-валидация недоверенного вывода", "5e1683ba"),
    ("vite.config.ts", "Config", "dev-прокси /deepseek (ключ server-side, вне бандла)", "5e1683ba"),
    ("src/pages/cargos/cargo-create/hooks/useAiFieldStatus.ts", "Хук (new)", "null-safe статус AI-полей (переиспользует шаги мастера)", "5e1683ba"),
    ("src/pages/auth/login/LoginPage.tsx", "UI", "Синхронный identity, Email на методе «код», onError по variables.data", "d597b2ff"),
    ("src/pages/auth/components/AuthIdentityInput.tsx", "UI (new)", "Общий переключатель Telegram/Email", "d597b2ff"),
    ("src/index.css", "CSS", "Флаг-кнопка телефона: overflow:hidden, снят min-height (~L1267)", "5e1683ba"),
    ("src/hooks/useVideoNoteRecorder.ts", "Хук", "rAF-таймер, canvas-preview, onerror, харднинг srcVideo", "fa72742b"),
    ("src/pages/chat/hooks/useHoldToRecordGesture.ts", "Хук (new)", "Удержание/свайп-блок/свайп-отмена (document-level pointer)", "fa72742b"),
    ("src/pages/admin/moderation/components/PassportModerationPanel.tsx", "UI (new)", "Паспортная модерация: Водители (/kyc/*) + Менеджеры (PATCH)", "d04332f8"),
    ("src/pages/admin/moderation-log/moderationLogUtils.ts", "Логика", "moderationLogType — фасет «тип модерации» в логе", "d04332f8"),
    ("public/locales/*.json ×15", "i18n", "Группа offerType (SIMPLE/COUNTER) — фикс «Simple»", "раб. дерево"),
    ("src/pages/offers/utils/offersPageHelpers.ts", "Логика", "offerTypeLabel (уже искал offerType.${KEY}) — кода не трогали", "—"),
]
for i, row in enumerate(files, start=2):
    for c, v in enumerate(row, start=1): put(ws3, i, c, v)

# ---------- Sheet 4: QA ----------
ws4 = wb.create_sheet("Проверки")
for i, w in enumerate([34, 18, 60], start=1):
    ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
for c, h in enumerate(["Проверка", "Результат", "Примечание"], start=1): put(ws4, 1, c, h)
style_header(ws4, 3)
qa = [
    ("pnpm build (tsc -b && vite build)", "✅ 0 ошибок", "exit 0 после каждой правки"),
    ("eslint --max-warnings 0", "✅ 0 предупреждений", "на всех затронутых файлах"),
    ("vitest (полный прогон)", "✅ 822/822", "на этапе AI-создания (15:20)"),
    ("Новые сьюты", "✅", "authEmail/LoginIdentity/OtpChannel, passportModeration/moderationLogType, cargoAi*, chatRecordingFormat, offersPageHelpers"),
    ("DeepSeek ключ", "✅ 200", "GET /models; прокси-заголовок вне бандла"),
    ("Локали", "✅ 15/15", "offerType + auth/AI ключи валидны во всех языках"),
    ("Авторизация (runtime)", "✅ Playwright", "публичные страницы логин/регистрация/сброс, light/dark, mobile overflow=0"),
    ("code-review (мульти-агент)", "✅", "AI-создание (5 фиксов), auth, паспортная модерация"),
    ("AI-создание (runtime)", "⚠ частично", "плагинг проверен живьём; полный сабмит за логином; бэкенд-прокси до прода"),
    ("Чат-рекордер / админ (runtime)", "⚠ на staging", "камера/HTTPS и админ-GUI за логином — недоступно для съёмки здесь"),
]
for i, row in enumerate(qa, start=2):
    for c, v in enumerate(row, start=1): put(ws4, i, c, v)

for ws_ in wb.worksheets:
    ws_.freeze_panes = "A2"

wb.save("report-07.07.26.xlsx")
print("saved report-07.07.26.xlsx")
