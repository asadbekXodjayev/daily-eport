# Builds report-18.07.26.xlsx — summary of the day's work: admin company tariffs, full-profile
# company registration, company overview (all fields), company team chat reuse, identity-based
# invitations + driver invite, dispatcher invitation bell, offers polling, chat call duration,
# WebRTC answer-path + camera/mic release. Run: python build-xlsx.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'report-18.07.26.xlsx')

BLUE = '1639C8'; DARK = '0F1D4F'; GREEN = '00B87A'; OKG = '0F9D58'
LINE = 'E6E9F0'; MUTEDBG = 'FAFBFE'
thin = Side(style='thin', color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)
top = Alignment(vertical='top')

wb = Workbook()

# -- Sheet 1: Changes --------------------------------------------------------------------
ws = wb.active
ws.title = 'Changes'
ws.sheet_view.showGridLines = False

title = ws.cell(row=1, column=1, value='Sarbon - Company tariffs / registration / overview / chat / invitations / web-calls - 18.07.2026')
title.font = Font(bold=True, size=15, color=DARK)
ws.cell(row=2, column=1,
        value='Admin company tariffs+limits. Full-profile company registration. Company overview (all ~30 fields + requisites + deletion banner). Company team chat reused over real /chat + calls. Identity-based employee invitations + in-app offers card + driver-invite mode + role dropdown fixes. Dispatcher invitation bell. Company offers polling. Chat call duration humanized. WebRTC answer-path SDP + camera/mic release. 90+ files / 3 commits + working-tree.'
        ).font = Font(size=10, color='5B6477', italic=True)
ws.merge_cells('A2:F2')

headers = ['#', 'Type', 'Area', 'Change', 'Endpoint / detail', 'Screenshot']
hr = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=BLUE)
    cell.alignment = Alignment(vertical='center', horizontal='left')
    cell.border = border

rows = [
    ['1', 'FEATURE', 'Admin / Tariffs',
     'Creator-only Tariffs tab (catalog list/create/edit/default) + per-company Assign-tariff modal: assigning copies the tariff limits into the company (existing companies unaffected by later edits); over-quota is a warning. Company limits change ONLY via tariff, not direct max_* PATCH. Pure tariffs.ts + tests',
     'POST /v1/admin/companies/:id/tariff {tariff_id}', '01-admin-tariff-create, 02-admin-assign-tariff'],
    ['2', 'FEATURE', 'Company / Registration',
     'Company create form expanded from 3 fields to a full profile (contacts, legal/director, bank details) - all optional; only name, type and INN (9-digit STIR / 14-digit PINFL) required. Create already returns a company-scoped token pair; optionals sent via a follow-up PATCH',
     'POST /v1/companies + PATCH', '03-company-create-profile'],
    ['3', 'FIX', 'Company / Overview',
     'Overview showed only ~8 of ~30 GET /v1/companies/:id fields. Added requisites card, owner (with "you" badge), tariff, updated_at; merged fleet+limits into one quota card (usage + limit, over-quota mark); all dates -> Tashkent; single empty placeholder. Redesigned deletion-request banner (amber Trash2 chip + withdraw). Pure companyOverview.ts + tests',
     'GET /v1/companies/:id, DELETE deletion-request', '04-company-overview, 05-company-overview-dark'],
    ['4', 'FEATURE', 'Company / Chat',
     'Company team chat = the real dispatcher ChatPage reused at /company/chat/:companyId (sidebar, media, voice/video notes, CALLS), made surface-aware via useChatIdentity + chatSurface.ts (company token under /company/*; /chat byte-identical). Cards switch-on-click via useCompanyOpen',
     '/company/chat/:companyId, /v1/chat/*', '06-company-chat'],
    ['5', 'FEATURE', 'Company / Invitations',
     'Moved employee invites from deprecated link/token to identity-based model: invite by phone -> 201 {id, status:pending} + in-app invitation_offer; invitee reads GET /company-users/offers, accepts by id. New "Invitations for you" card (Accept/Decline) on company home. must_register_first (404) / already_member (409) handled',
     'POST /v1/invitations/:id/accept|decline, GET /company-users/offers', '07-invitation-offers-card'],
    ['6', 'FIX', 'Company / Role dropdown',
     'companyRoleI18nKey normalizes casing (ALL-CAPS / PascalCase / SNAKE_CASE all resolve) so labels no longer render "Cargomanager"; parseInvitableCompanyRoleOptions keeps only company_user_roles employee roles and always drops Owner/Carrier/Driver; SHIPPER type-gate hides driver/fleet roles',
     'GET /v1/reference/company', '08-invite-role-dropdown, 09-invite-role-gating-shipper'],
    ['7', 'FEATURE', 'Company / Driver invite',
     'Employee|Driver toggle in the invite modal: Driver mode hides the role dropdown and submits phone-only to the separate driver endpoint (permission fleet.driver.manage, all 3 company types). Driver accepts in the driver app. buildCreateDriverInvitationPayload + driverInvitationErrorCode (+7 tests)',
     'POST /v1/companies/:id/driver-invitations {phone}', '10-driver-invite'],
    ['8', 'FEATURE', 'Dispatcher / Notifications',
     'Dispatcher notification bell now renders the invitation_offer type (was invisible to invited CM/DM): "Company invitation" tag, Building2 icon, localized "{inviter} invited you to join {company} as {role}", "Open invitation" CTA -> /:lang/company, explicit SSE handler + OS notification',
     'invitation_offer (SSE + trip-notifications)', '11-dispatcher-invitation-bell'],
    ['9', 'FIX', 'Company / Offers',
     'Company Offers (pending-approval) and per-cargo offers only revalidated on the 60s staleTime; company surface has no realtime bridge. Added polling + refetch-on-focus (staleTime 15s, refetchInterval 20s, refetchOnWindowFocus true) so a bid landing while a manager sits on the page appears',
     'GET /v1/companies/:id/offers/pending-approval', '12-offers-polling'],
    ['10', 'FIX', 'Chat / Call duration',
     'Conversation-list preview showed a call duration as raw seconds "(73s)" instead of "1 min 13 sec"; preview now rebuilt client-side (formatChatCallEventPreview) to match the thread bubble and follow current language. Sub-minute calls drop the "0 min" prefix -> "22 sec" (new durationSecondsLabel in all 15 locales)',
     'ConversationItem, chatCallEvents.ts', '13-chat-call-duration-sidebar, 14-chat-call-duration-thread'],
    ['11', 'FIX', 'Web calls / WebRTC',
     'Answer (callee) path mobile->web: web ANSWER was a=recvonly audio + a=msid:- video (freeze/silence); forceVideoSendrecv now setStreams (real msid) + new forceAudioSendrecv/findAudioTransceiver re-assert both m-lines before createAnswer, so the 300kbps/20fps cap takes effect. Also: camera/mic now released unconditionally on endCall (was skipped on 409/error) + reconcile guard covers ended',
     'webrtcTuning.ts, GlobalAudioCallModal.tsx', '(protocol/media - no UI; verify-webrtc-answer-sdp.mjs)'],
]
type_fill = {'FIX': OKG, 'FEATURE': GREEN, 'UI': '7C3AED'}
r = hr + 1
for row in rows:
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = wrap if c in (4, 5, 6) else top
        cell.font = Font(size=10)
        if c == 2:
            cell.font = Font(size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=type_fill.get(val, '64748B'))
            cell.alignment = Alignment(vertical='center', horizontal='center')
    r += 1

widths = [5, 10, 24, 56, 40, 28]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for rr in range(hr + 1, r):
    ws.row_dimensions[rr].height = 84

# -- Sheet 2: Verification ---------------------------------------------------------------
ws2 = wb.create_sheet('Verification')
ws2.sheet_view.showGridLines = False
ws2.cell(row=1, column=1, value='Verification gates').font = Font(bold=True, size=14, color=DARK)
gates = [
    ['Gate', 'Result'],
    ['tsc -b (full project type-check)', 'exit 0'],
    ['vitest run', '1950 / 1950'],
    ['eslint --max-warnings 0 + prettier (changed files)', 'clean'],
    ['Locale parity', '15/15 (company.* / admin.* / tripNotifications.* en/uz/ru + placeholders)'],
    ['Runtime verification (per fix)', '14 new qa/verify-*.mjs, 0 console errors'],
    ['WebRTC answer-path SDP (verify-webrtc-answer-sdp.mjs)', 'both m-lines sendrecv + real msid'],
    ['Adversarial / code-review follow-ups', 'findings addressed'],
    ['Files touched', '90+ (3 commits + working-tree)'],
]
gr = 3
for i, (a, b) in enumerate(gates):
    ca = ws2.cell(row=gr + i, column=1, value=a)
    cb = ws2.cell(row=gr + i, column=2, value=b)
    for cell in (ca, cb):
        cell.border = border
        cell.alignment = top
    if i == 0:
        for cell in (ca, cb):
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=BLUE)
    else:
        ca.font = Font(size=10)
        cb.font = Font(size=10, bold=True, color=OKG)
ws2.column_dimensions['A'].width = 56
ws2.column_dimensions['B'].width = 52

wb.save(OUT)
print('wrote', OUT)
