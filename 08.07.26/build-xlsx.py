# Builds report-08.07.26.xlsx — PM summary of 08.07.2026 Sarbon Frontend work (4 sheets).
# Assembled from the 9 live screenshots captured 08.07 (admin Users-registry Insights section).
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BLUE = "1639C8"; SOFT = "EEF1FF"; INK = "16181D"; GREEN = "0F9D58"; MUTED = "5B6477"
hdr_fill = PatternFill("solid", fgColor=BLUE)
sub_fill = PatternFill("solid", fgColor=SOFT)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
lbl_font = Font(bold=True, color=INK)
wrap = Alignment(wrap_text=True, vertical="top")
thin = Side(style="thin", color="D9DEE8")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = openpyxl.Workbook()

def style_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = wrap; cell.border = border

def put(ws, r, c, v, font=None, fill=None, al=wrap):
    cell = ws.cell(row=r, column=c, value=v)
    if font: cell.font = font
    if fill: cell.fill = fill
    cell.alignment = al; cell.border = border
    return cell

# ---------- Sheet 1: Summary ----------
ws = wb.active; ws.title = "Сводка"
ws.column_dimensions['A'].width = 34; ws.column_dimensions['B'].width = 78
put(ws, 1, 1, "Отчёт о работе · 08.07.2026 · Sarbon Frontend", hdr_font, hdr_fill)
put(ws, 1, 2, "Фокус: секция Insights / Trend реестра пользователей в админке — 3 ролевые вкладки, KPI-карточки, контекстные алерты, тренд с метрикой × гранулярностью", hdr_font, hdr_fill)
rows = [
    ("Дата", "08.07.2026"),
    ("Проект", "sarbon-frontend · React 19 · Vite · Ant Design 6"),
    ("Ветка", "staging"),
    ("Область", "Админ · Users registry (/admin/users) · секция Insights / Trend"),
    ("Главное", "Единый аналитический экран по 3 ролям: водители · карго-менеджеры · менеджеры водителей"),
    ("Ролевые вкладки Insights", "3 · Drivers, Cargo managers, Driver managers"),
    ("KPI-карточек", "13 (по всем ролям)"),
    ("Умных баннеров-алертов", "2 · KYC queue needs attention · Fleet coverage gap detected"),
    ("Секция Trend", "селектор метрики × гранулярность День/Неделя/Месяц × диапазон дат × график"),
    ("Ключевая корректность", "снапшот (DAU) vs временной ряд («Новые …»): DAU без выдуманного тренда"),
    ("KYC-контроль (водители)", "Identity checks pending (approval 86.4%) + stuck 7+ days (oldest 15.1d)"),
    ("Пробел покрытия (менеджеры водителей)", "Managers with no drivers 9 (60% of all DMs) · Average fleet size 2.5"),
    ("Пустые состояния (карго-менеджеры)", "Active cargos / No offer in 24h → корректное «No data»"),
    ("Скриншотов-доказательств", "9 (живой Playwright с авторизацией под администратором)"),
    ("JS-исключений в прогоне", "0"),
    ("Скрипты захвата", "capture-trend.cjs · recapture-managers.cjs"),
    ("Примечание", "Отчёт (html + Excel) собран позже из скриншотов 08.07; на день был только шаг захвата"),
    ("Следующий рабочий день", "09.07 — тренд-график линия → столбцы (AdminLine → AdminTrendBars)"),
]
for i, (k, v) in enumerate(rows, start=2):
    put(ws, i, 1, k, lbl_font, sub_fill); put(ws, i, 2, v)

# ---------- Sheet 2: Changes ----------
ws2 = wb.create_sheet("Изменения")
widths = [4, 14, 34, 46, 52, 26]
for i, w in enumerate(widths, start=1):
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
headers = ["№", "Тип", "Область", "Что было / причина", "Что сделано", "Роль / поверхность"]
for c, h in enumerate(headers, start=1): put(ws2, 1, c, h)
style_header(ws2, len(headers))
changes = [
    ("feature", "Реестр пользователей — ролевые вкладки Insights",
     "Показатели по типам пользователей были разрознены; не было единого экрана «здоровья» аудитории.",
     "Сегмент Drivers / Cargo managers / Driver managers над общим макетом Insights; смена вкладки перезагружает KPI-карточки, алерты и источник тренда для роли.",
     "Админ · /admin/users"),
    ("feature", "KPI-карточки водителей + KYC-контроль",
     "Не было наглядного контроля очереди проверки личности (KYC) и «протухших» заявок.",
     "Online/DAU, Weekly active, New this week (Today/Month), Identity checks pending (approval rate) и stuck 7+ days; красный баннер «KYC queue needs attention».",
     "Админ · вкладка Drivers"),
    ("feature", "KPI карго-менеджеров + пустые состояния",
     "Нужен срез активности менеджеров грузов и «спящих» позиций.",
     "Online/DAU (% of total), Daily active, New this week, Active cargos, No offer in 24h; где метрики нет — корректное «No data» вместо нуля/ошибки.",
     "Админ · вкладка Cargo managers"),
    ("feature", "KPI менеджеров водителей + алерт покрытия флота",
     "Не было видно менеджеров без закреплённых водителей (пробел покрытия).",
     "Online/DAU, Daily active, New this week, Average fleet size, Managers with no drivers (доля от всех DM); красный баннер «Fleet coverage gap detected».",
     "Админ · вкладка Driver managers"),
    ("feature", "Секция Trend — метрика × гранулярность × диапазон",
     "Показатели были только «на сейчас», без динамики во времени.",
     "Селектор метрики, сегмент День/Неделя/Месяц, пикер диапазона дат, линейный график (recharts); «Новые …» бакетируются на фронте в ряд по периодам.",
     "Админ · все три вкладки"),
    ("correctness", "Честный снапшот vs временной ряд",
     "Точечную метрику (DAU) легко ошибочно нарисовать «трендом», хотя истории нет — это вводит в заблуждение.",
     "Для снапшот-метрик контролы приглушаются, рядом тег «Point-in-time — no historical trend»; график рисует ровную линию текущего значения, без выдуманной динамики.",
     "Админ · Trend"),
]
for i, row in enumerate(changes, start=2):
    put(ws2, i, 1, i - 1)
    for c, v in enumerate(row, start=2):
        put(ws2, i, c, v)

# ---------- Sheet 3: Screenshots ----------
ws3 = wb.create_sheet("Скриншоты")
widths = [6, 40, 66]
for i, w in enumerate(widths, start=1):
    ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
for c, h in enumerate(["№", "Файл", "Что на кадре"], start=1): put(ws3, 1, c, h)
style_header(ws3, 3)
shots = [
    ("01", "01-users-drivers-overview.png", "Вкладка Drivers: KPI-карточки, баннер «KYC queue needs attention», секция Trend."),
    ("02", "02-trend-new-drivers-day.png", "Тренд «Новые водители» · День — реальный ряд (бакетирование на фронте), тултип по точке."),
    ("03", "03-trend-new-drivers-week.png", "Тренд «Новые водители» · Неделя — перебакетирование ряда по неделям."),
    ("04", "04-trend-new-drivers-month.png", "Тренд «Новые водители» · Месяц — агрегирование ряда по месяцам."),
    ("05", "05-trend-dau-snapshot.png", "DAU — снапшот: контролы приглушены, тег «Point-in-time — no historical trend»."),
    ("06", "06-cargo-managers-overview.png", "Вкладка Cargo managers: KPI + корректные пустые состояния «No data»."),
    ("07", "07-trend-new-cargo-managers.png", "Тренд «Новые карго-менеджеры за неделю» — ряд по периодам для роли."),
    ("08", "08-driver-managers-overview.png", "Вкладка Driver managers: KPI + баннер «Fleet coverage gap detected» (9 DM без флота)."),
    ("09", "09-trend-new-driver-managers.png", "Тренд «Новые менеджеры водителей за неделю» — ряд по периодам для роли."),
]
for i, row in enumerate(shots, start=2):
    for c, v in enumerate(row, start=1):
        put(ws3, i, c, v)

# ---------- Sheet 4: Checks ----------
ws4 = wb.create_sheet("Проверки")
ws4.column_dimensions['A'].width = 42; ws4.column_dimensions['B'].width = 62
for c, h in enumerate(["Проверка", "Результат"], start=1): put(ws4, 1, c, h)
style_header(ws4, 2)
checks = [
    ("Живой прогон", "Playwright/Chromium против живого бэкенда, авторизация под администратором"),
    ("JS-исключения в рантайме", "0 в прогоне"),
    ("Вкладка Drivers", "KPI + KYC-алерт + тренд (Day/Week/Month) + DAU-снапшот — снято"),
    ("Вкладка Cargo managers", "KPI (+ «No data») + тренд «Новые за неделю» — снято"),
    ("Вкладка Driver managers", "KPI + «Fleet coverage gap» + тренд «Новые за неделю» — снято"),
    ("Снапшот vs ряд", "DAU помечен «Point-in-time», контролы приглушены — корректно"),
    ("Скриншотов отчёта", "9 (живой Playwright с авторизацией)"),
    ("Примечание", "Отчёт собран позже из скриншотов 08.07; на день был только шаг захвата"),
]
for i, (k, v) in enumerate(checks, start=2):
    put(ws4, i, 1, k, lbl_font, sub_fill); put(ws4, i, 2, v)

out = "C:/Users/hp/Desktop/report/daily-report/08.07.26/report-08.07.26.xlsx"
wb.save(out)
print("saved", out)
