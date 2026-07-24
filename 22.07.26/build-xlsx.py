# Builds report-22.07.26.xlsx — the day's work: company cargo create/edit (a surface seam over the
# dispatcher wizard) + paged list + board drill-through + offers driver column, in-place invitation
# accept/decline, server-read role permissions, four company-auth fixes, the production push and the
# restored public entry points, and a post-merge architecture audit that found a cross-account cache leak.
# Run: python build-xlsx.py
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, 'report-22.07.26.xlsx')

BLUE = '1639C8'; DARK = '0F1D4F'; GREEN = '00B87A'; OKG = '0F9D58'
LINE = 'E6E9F0'
thin = Side(style='thin', color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(vertical='top', wrap_text=True)
top = Alignment(vertical='top')

wb = Workbook()

# -- Sheet 1: Changes --------------------------------------------------------------------
ws = wb.active
ws.title = 'Changes'
ws.sheet_view.showGridLines = False

title = ws.cell(row=1, column=1,
                value='Sarbon - Company cargo create/edit + paged list / in-place invitation accept / server-read role permissions / production push / cross-account cache leak - 22.07.2026')
title.font = Font(bold=True, size=15, color=DARK)
ws.cell(row=2, column=1,
        value='Five commits and two merges (PR #299, #300) plus substantial working-tree work. Morning: finishing yesterday\'s company-surface work as the backend delivered the endpoints it had promised. Afternoon: the staging -> main production push, with four public company entry points temporarily commented out and restored afterwards. Post-merge: an architecture audit that found a cross-account cache leak, and the day\'s largest task - company cargo create/edit did not exist at all and the cargo list silently showed only its first page. 58 files in src/ and locales (8 new modules, 15 locale files), 9 bug.md entries.'
        ).font = Font(size=10, color='5B6477', italic=True)
ws.merge_cells('A2:F2')

headers = ['#', 'Type', 'Area', 'Change', 'Endpoint / detail', 'Screenshot']
hr = 4
for c, h in enumerate(headers, start=1):
    cell = ws.cell(row=hr, column=c, value=h)
    cell.font = Font(bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=BLUE)
    cell.alignment = Alignment(vertical='center', horizontal='left')
    cell.border = border

rows = [
    ['1', 'FEATURE', 'Company workspace / Cargo',
     'A company could view cargo but could not CREATE one at all. The repo\'s only cargo form is hardwired to the dispatcher identity (useGet/usePost/usePut + ENDPOINTS_DISPATCHERS_V1) with no seam for a second surface - and reusing it naively is WORSE than not reusing it: none of its ~14 requests are in isPublicAuthUrl, so under /company/* they resolve to authMode dispatcher, and a 401 there ends in a hard location.replace to the DISPATCHER login, ejecting the company user from their own workspace. Built a surface seam instead of a fork (mirroring the chat-reuse precedent): cargoFormSurface.ts describes the hosting surface (auth mode, create/update URL, verb, invalidation keys, navigate-after-save, visibility support) and CargoFormSurfaceContext carries it to the wizard leaves. The dispatcher default leaves authMode undefined, so its requests stay byte-identical',
     'POST /v1/companies/:id/cargo, PATCH /v1/companies/:id/cargo/:cargoId',
     '02-company-cargo-create-drawer, 04-company-cargo-edit-drawer'],

    ['2', 'FEATURE', 'Cargo / Search visibility',
     'Cargo creators now choose who sees a posting: "all" (SEARCHING_ALL - every Sarbon driver, including other companies\') vs "company" (SEARCHING_COMPANY - only that company\'s own drivers and driver managers), sent as CargoCreateRequest.desired_search_visibility. Offered to anyone posting ON BEHALF OF A COMPANY: the company workspace, and /cargo-create when a dispatcher has context-switched into a company. A freelance dispatcher never sees it (the backend force-overrides to "all" for an unscoped dispatcher). Deliberately NOT offered on edit: CargoUpdateRequest has no such field and the only endpoint that rewrites visibility afterwards is a legacy dispatcher route that rejects a company token with a bare 400 (reproduced on staging). An earlier iteration wired it and had to be removed - besides never succeeding, it seeded a hidden desired_search_visibility into the form on every edit, so saving an unrelated field on a PENDING_MODERATION cargo fired a spurious status transition. BACKEND ASK: a company-scoped endpoint to change an existing cargo\'s visibility',
     'CargoCreateRequest.desired_search_visibility', '03-company-cargo-visibility-control'],

    ['3', 'FIX', 'Company workspace / Cargo list',
     'The list sent NO query params at all, so a company with 45 postings saw the server\'s default first 20 presented as the complete list. Now sends page/limit/status/sort/assigned_manager_id with a real pager and facets (companyCargoList.ts). Verified on the wire: ?limit=20&page=1&sort=created_at:desc',
     'GET /v1/companies/:id/cargo', '01-company-cargo-list-paged, 12-company-cargo-pager'],

    ['4', 'FIX', 'Company workspace / Board + Offers',
     'Board cards were inert - a trip card now deep-links to that trip\'s detail drawer via a one-shot ?trip= param that is stripped after being consumed. companyOffers.ts never read the nested carrier.name the spec documents, so the Driver column rendered empty for every row shaped as documented. CargoForm never passed cargoId to Step1CargoInfo, so its edit path skipped photo hydration (caught by the new browser verifier). board.cardStatus.* was missing from all 15 locales and always fell back to a raw enum string. CargoForm used antd\'s static message instead of App.useApp()',
     'GET /v1/companies/:id/board, .../offers/pending-approval',
     '05-company-board, 10-company-board-trip-drilldown, 07-company-offers-driver-column'],

    ['5', 'UI', 'Company workspace / Operations visual sweep',
     'Screenshot sweep of 4 tabs x 3 viewports x 2 themes + drawer states = 32 shots. Six defects, all found by eye rather than by the automated checks. (1) HIGH: the "Actions" header in Trips rendered one letter per line, inflating the header to ~190px - declared column widths summed above scroll.x, so AntD compressed every column and a Russian word with no break opportunity wrapped per character; fixed with a raised scroll.x plus a withNowrapHeaders() helper, since widths alone are locale-fragile. (2) role tag overflowed into the price column. (3) the price was the field that truncated on board cards - now the timestamp wraps instead. (4) identical cards truncated routes differently because a neighbour\'s longer status tag stole width - status tag moved above the route. (5) two rows formatted the same currency differently - the fallback now routes through the app-wide formatPrice. (6) the cargo action group ran 45px past the card edge at 390px. Result: overflowing shots 6 -> 0, console errors 0',
     'qa/shoot-company-operations.mjs, qa/reports/qa-report-2026-07-22-company-operations.md',
     '08-company-trips-table, 05-company-board, 07-company-offers-driver-column'],

    ['6', 'FEATURE', 'Offers / Company invitation',
     'The backend shipped the direct-accept path, so yesterday\'s OTP detour is legacy. POST /v1/dispatchers/invitations/{id}/accept accepts the invitation in place with the DISPATCHER token, auto-provisions a company_user identity by phone (dispatcher profile snapshot, role written neutrally - not OWNER) and returns a company_user token pair; phone must match the dispatcher (else 403 invitation_not_yours). The card was rewired from a single "Open invitation" button to in-place Accept + Decline: accept -> clearCompanySessionQueries + setAuth (auto-login) -> invalidate the feed -> navigate into the joined company\'s workspace. A 200 with no usable tokens fails loudly instead of half-persisting a session',
     'POST /v1/dispatchers/invitations/{id}/{accept,decline}', '20-dispatcher-invitation-accept-decline'],

    ['7', 'FIX', 'Offers / Invitation accept (adversarial review)',
     'A 3-lens review of the above caught one real MEDIUM: parseDispatcherInvitationAccept accepted a token pair whose expires_in / refresh_expires_in were missing or non-numeric by defaulting them to 0 - and since the token store writes expiry as (now + lifetime), 0 becomes an already-past epoch rather than the documented "never expires" sentinel. The first company request would refresh -> fail -> clearAuth and bounce the user to the company login immediately after the success toast. The parser now fails closed',
     'src/pages/offers/components/dispatcherInvitationAccept.ts', '(logic - 9 parser tests)'],

    ['8', 'FEATURE', 'Company workspace / Roles',
     'The role-permission read endpoint the frontend requested yesterday shipped: GET /v1/companies/{companyId}/roles/{role}/permissions. Wired it and made the QUERY CACHE the single source of truth - new parseRolePermissionOverlay parses the GET into a permission_key -> allowed map (tolerant of envelope/bare/array, drops malformed rows, non-boolean allowed never coerced); save and per-key revert patch the cache optimistically via setQueryData (lossless round-trip with the existing payload builder) then invalidate to reconcile. Yesterday\'s session-only "applied" state and its effect-seed were removed entirely. Saved state now survives a page refresh, a tab reopen, a role switch-and-back, and is visible to another admin',
     'GET /v1/companies/:id/roles/:role/permissions', '09-company-roles-server-overlay'],

    ['9', 'FIX', 'Company workspace / Roles (adversarial review)',
     'The overlay GET\'s loading and error states were never surfaced, so a slow or failed read rendered every permission as "Not set" - indistinguishable from a role with no overrides, and editable from that false-blank baseline. The read now shares the section\'s error+retry block (retry refetches both, since both are roles.view-gated and the overlay is a prerequisite for correct display) and the toggles are disabled until it loads',
     'RolesSection.tsx', '09-company-roles-server-overlay'],

    ['10', 'FIX', 'Company auth / Conflict handling',
     'isCompanyAlreadyRegistered (registration complete) and isCompanyInnTakenError (create company) each asserted nothing but the HTTP status, so ANY 409 was relabelled: on complete it toasted "you already have an account" and navigated to login with replace:true, discarding the typed name/password and the session_id in the URL; on create it pinned "INN already taken" onto a correct INN, an error the user could not clear by editing the field. New isExpectedConflict(error, status, reasons) + apiErrorReason: the status still matches on its own when the body names no reason (the documented shape), but a 409 that DOES name a different reason falls through to a plain in-place error',
     'companyAuthError.ts, companyRegisterConflict.ts, companyCreate.ts', '(logic - covered by companyAuthError tests)'],

    ['11', 'FIX', 'Auth / OTP cooldown',
     'sarbon.otpFailedIdentifier is the single app-wide marker behind the 60-second repeated-failure throttle, and only the dispatcher pages write it. When a company-SCOPED cooldown expired, the hook cleared it anyway - so the next failed send on the dispatcher login counted as a first failure and skipped the throttle entirely. Only the global (unscoped) instance may clear it now. Known remainder, documented in the hook rather than worked around: switching the field to a different identifier and back can render one second of a stale countdown before the tick corrects it - both clean fixes are rejected by react-hooks/purity and react-hooks/set-state-in-effect',
     'src/pages/auth/hooks/useOtpCooldown.ts', '(logic)'],

    ['12', 'FIX', 'Company auth / Registration complete',
     'POST /registration/complete can answer with a body naming the offending field (data.fields), but onError tested the status branches FIRST and those branches return after navigating - so a 409 carrying a password field error discarded the message and threw the user off a form they could have corrected in place. applyApiFieldErrors now runs first; the redirects are unaffected because their responses carry no fields',
     'CompanyCompletePage.tsx', '(logic)'],

    ['13', 'FIX', 'Company auth / Error text',
     'companyAuthErrorText replaces a machine token with the caller\'s localized fallback so raw service text never reaches the user, but codes without a separator - unauthorized, forbidden, conflict - passed the filter and rendered verbatim in a toast. The separator rule is deliberate (Uzbek is the default language and is lowercase Latin, so a one-word message like "bloklangan" must reach the user), so the fix is a CLOSED allowlist of single-token codes rather than widening the rule to "any English word"',
     'companyAuthError.ts', '(logic)'],

    ['14', 'REFACTOR', 'Company auth / Dead identity barrier',
     'companyRegisterConflict.ts modelled a backend barrier that no longer exists: a 409 with data.reason phone_belongs_to_dispatcher / phone_belongs_to_driver. Two pages branched on it, two i18n keys x 15 locales carried its copy, a test suite asserted it. The live spec is explicit on both halves - purpose=register "no longer influences anything", and a phone owned by a dispatcher OR a driver "is no longer rejected"; phone_belongs_to_* appears 0 times. Removed, and CLAUDE.md corrected (it still documented the barrier as live, which is what caused a finding to be over-rated the day before)',
     'companyRegisterConflict.ts, 15 locales, CLAUDE.md', '(no UI)'],

    ['15', 'RELEASE', 'Production push + entry points',
     'staging -> main (PR #300). Before the push the four public company entry points were COMMENTED OUT rather than deleted (per the explicit instruction "hide them, do not kill them"), each with an uncomment-to-restore marker: the footer link, the guest header item, and the cross-links on the dispatcher login and register pages. /company/* and /accept-invite stayed reachable by direct URL, as did the offers invitation card, the hired-dispatcher context switcher and admin moderation. After the push all four were restored to their pre-hide versions; tsc -b confirms they still typecheck against the merged OTP-cooldown refactor',
     'Footer.tsx, UserProfile.tsx, LoginPage.tsx, RegisterPage.tsx',
     '30-footer-for-companies, 31-header-for-companies, 32-login-cross-identity-hint'],

    ['16', 'FIX', 'Company session / Cache purge (HIGH)',
     'Found by a post-merge architecture audit: the merge fixed the auth/create/profile/admin layers but shipped only HALF of its own review\'s finding - the login/logout cache scope was left open. clearCompanySessionQueries, called on every session transition (login / OTP-verify / registration-complete / logout), removed only 4 account-level keys. The ~25 company-SCOPED workspace keys (trips, finance, users, documents, drivers, invitations, audit, board, ...), keyed [key, companyId], survived the transition under the 5-minute gcTime. FAILURE SCENARIO: two employees of the same company on a shared device - one logs out, the other logs in to the same companyId, and for the first ~60s sees the first employee\'s cached, differently-permissioned data. Fixed by purging on MEMBERSHIP - removeQueries({predicate}) over the full QUERY_KEYS_COMPANY_V1 set - so a newly-added company endpoint cannot silently escape the purge',
     'src/utils/companySessionCache.ts', '(cache-level - covered by a new unit test)'],

    ['17', 'ASSESS', 'Architecture audit (post-merge)',
     'Compared the company-surface-fixes merge against the pre-merge base across 6 areas plus a structural audit. VERDICT: a genuine improvement, 0 regressions. It fixed several real base bugs (an inverted 409-reason null-check misclassifying any unknown conflict; an unparseable 200 stranding the user on the create form; an is_current fetch-gate producing an un-retryable 403; a global cross-surface OTP cooldown; OTP attempts burned on transient errors) and improved logic/view separation by extracting seven pure modules each with a co-located test. No new structural holes (knip\'s "36 unused files" are all standalone qa/*.mjs scripts, a config artifact). One HIGH open hole found and fixed - row 16. Deferred, reasons unchanged: god-components (FinanceSection 1233 LOC - needs a GUI baseline first), dead scaffolding pending the backend, and the qa/verify-*.mjs portability issue',
     'ARCHITECTURE.md change log', '(no UI)'],
]
type_fill = {'FIX': OKG, 'FEATURE': GREEN, 'UI': '7C3AED', 'REFACTOR': '64748B', 'RELEASE': '0EA5E9', 'ASSESS': 'B8860B'}
r = hr + 1
for row in rows:
    for c, val in enumerate(row, start=1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        cell.alignment = wrap if c in (4, 5, 6) else top
        cell.font = Font(size=10)
        if c == 2:
            cell.font = Font(size=9, bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=type_fill.get(val, '64748B'))
            cell.alignment = Alignment(vertical='center', horizontal='center')
    r += 1

widths = [5, 11, 30, 64, 38, 34]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for rr in range(hr + 1, r):
    ws.row_dimensions[rr].height = 100

# -- Sheet 2: Verification ---------------------------------------------------------------
ws2 = wb.create_sheet('Verification')
ws2.sheet_view.showGridLines = False
ws2.cell(row=1, column=1, value='Verification gates - 22.07.2026').font = Font(bold=True, size=14, color=DARK)
gates = [
    ['Gate', 'Result'],
    ['pnpm build (tsc -b + vite)', 'exit 0'],
    ['eslint . --max-warnings 0', 'clean'],
    ['vitest (full run, end of day)', '192 files / 2161 tests'],
    ['qa/verify-company-cargo-create.mjs (new)', '22/22 with 0 console errors'],
    ['qa/verify-company-workspace.mjs', '18/18 sections, 0 raw i18n keys'],
    ['qa/verify-company-invite-roles.mjs', 'exit 0'],
    ['Operations screenshot sweep (32 shots)', 'overflowing shots 6 -> 0, console/JS errors 0'],
    ['Locale parity', '15/15 (+34 keys per locale in the cargo task, none removed)'],
    ['Dispatcher /cargo-create left untouched', 'NOT exercised under the stubbed harness - the untouched my-cargos page renders identically blank there, so a harness limit rather than a regression; the no-change guarantee rests on the type gate, lint, and unit tests asserting the dispatcher surface keeps authMode undefined, PUT /api/cargo/:id, and no desired_search_visibility in its payload'],
    ['Changing an existing cargo\'s visibility', 'NOT POSSIBLE - needs a company-scoped endpoint (backend ask filed)'],
    ['Live end-to-end invitation accept', 'NOT exercised - no test invitation on staging; the server remains the source of truth for the phone match and guards'],
    ['Files touched', '58 in src/ and locales (8 new modules, 15 locale files), 5 commits + 2 merges + working tree'],
    ['bug.md entries', '9 for the day'],
]
gr = 3
for i, (a, b) in enumerate(gates):
    ca = ws2.cell(row=gr + i, column=1, value=a)
    cb = ws2.cell(row=gr + i, column=2, value=b)
    for cell in (ca, cb):
        cell.border = border
        cell.alignment = wrap if i > 0 else top
    if i == 0:
        for cell in (ca, cb):
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=BLUE)
    else:
        ca.font = Font(size=10)
        bad = ('NOT ' in b) or ('POSSIBLE' in b)
        cb.font = Font(size=10, bold=True, color='E8590C' if bad else OKG)
ws2.column_dimensions['A'].width = 46
ws2.column_dimensions['B'].width = 92
for rr in range(gr, gr + len(gates)):
    ws2.row_dimensions[rr].height = 42

# -- Sheet 3: Screenshots -----------------------------------------------------------------
ws3 = wb.create_sheet('Screenshots')
ws3.sheet_view.showGridLines = False
ws3.cell(row=1, column=1, value='Screenshots - all captured with Playwright against the running app').font = Font(bold=True, size=14, color=DARK)
ws3.cell(row=2, column=1,
         value='Company workspace and offers-feed screens run on a seeded session with stubbed staging responses - the same harness the change was verified with (qa/verify-company-cargo-create.mjs, qa/shoot-company-operations.mjs), because the owner surface has no shareable test account and there is no live pending invitation on staging. The offers feed was entered under a real dispatcher session with only the feed response stubbed. The public screens (footer, header, login) use no stubs at all.'
         ).font = Font(size=10, color='5B6477', italic=True)
ws3.merge_cells('A2:C2')
shots = [
    ['File', 'Shows', 'Mode'],
    ['01-company-cargo-list-paged.png', 'Cargo tab - status/curator facets, sort, "Add cargo"', 'seeded + stubbed'],
    ['12-company-cargo-pager.png', 'Pager for 45 rows at 20 per page (the list used to show only page 1)', 'seeded + stubbed'],
    ['02-company-cargo-create-drawer.png', 'The four-step wizard opened as a drawer inside the company workspace', 'seeded + stubbed'],
    ['03-company-cargo-visibility-control.png', '"Who will see this cargo": all drivers vs only my company', 'seeded + stubbed'],
    ['04-company-cargo-edit-drawer.png', 'Edit drawer - hydrated from the cargo, and no visibility control', 'seeded + stubbed'],
    ['05-company-board.png', 'Board - status tag above the route, price and timestamp on separate lines', 'seeded + stubbed'],
    ['10-company-board-trip-drilldown.png', 'Board card click -> Trips tab with that trip\'s detail drawer open', 'seeded + stubbed'],
    ['07-company-offers-driver-column.png', 'Driver column populated from the nested carrier.name; both prices formatted alike', 'seeded + stubbed'],
    ['08-company-trips-table.png', 'Trips table - headers on one line, role tag inside its own column', 'seeded + stubbed'],
    ['09-company-roles-server-overlay.png', 'Role permissions AFTER a page reload, read back from the new GET', 'seeded + stubbed'],
    ['20-dispatcher-invitation-accept-decline.png', 'Company invitation with in-place Accept / Decline in the dispatcher offers feed', 'live dispatcher session + stubbed feed'],
    ['30-footer-for-companies.png', 'Footer "For companies" link restored after the production push', 'live, no stubs'],
    ['31-header-for-companies.png', 'Guest header "For companies" entry restored', 'live, no stubs'],
    ['32-login-cross-identity-hint.png', 'Dispatcher login company cross-link restored', 'live, no stubs'],
]
sr = 4
for i, row in enumerate(shots):
    for c, val in enumerate(row, start=1):
        cell = ws3.cell(row=sr + i, column=c, value=val)
        cell.border = border
        cell.alignment = wrap
        if i == 0:
            cell.font = Font(bold=True, size=10, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor=BLUE)
        else:
            cell.font = Font(size=10)
ws3.column_dimensions['A'].width = 46
ws3.column_dimensions['B'].width = 84
ws3.column_dimensions['C'].width = 36
for rr in range(sr, sr + len(shots)):
    ws3.row_dimensions[rr].height = 30

wb.save(OUT)
print('wrote', OUT)
