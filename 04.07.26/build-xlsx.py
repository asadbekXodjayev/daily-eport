# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:\Users\hp\Desktop\report\daily-report\04.07.26\report-04.07.26.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1639C8")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="16181D")
SUB_FONT = Font(bold=False, size=10.5, color="5b6477")
CELL_FONT = Font(bold=False, size=10.5, color="16181D")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CTR = Alignment(wrap_text=True, vertical="center")
THIN = Side(style="thin", color="E6E9F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

# ---------- Sheet 1: Сводка ----------
ws = wb.active
ws.title = "Сводка"
ws.column_dimensions["A"].width = 52
ws.column_dimensions["B"].width = 3
ws.column_dimensions["C"].width = 36
ws["A1"] = "Отчёт о работе · 04.07.2026 · Sarbon Frontend"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Фокус: видеозвонок (Telegram-редизайн + исправления), видео-кружки, чистые названия городов, адаптив админ-консоли, карточка водителя"; ws["A2"].font = SUB_FONT
rows = [
    ("04.07.2026", "Дата"),
    ("sarbon-frontend  ·  React 19 · Vite · Ant Design 6", "Проект"),
    ("staging", "Ветка / окружение"),
    ("48c61aa5 · a5d6b315 · 37ceaf43 · 285fbfd6 · 9b17a6b8", "Коммиты за день"),
    ("~11", "Областей изменений"),
    ("6", "Функций / переработок"),
    ("8", "Дефектов закрыто"),
    ("~6", "Новых файлов"),
    ("~11", "Наборов тестов (новых / расширенных)"),
    ("718 / 718", "Юнит-тестов зелёных (vitest)"),
    ("15", "Локалей (новые ключи звонка / кружков)"),
    ("0", "Ошибок tsc -b / pnpm build"),
    ("0", "Предупреждений ESLint (затронутые файлы)"),
]
r = 4
for val, label in rows:
    ws.cell(row=r, column=1, value=val).font = CELL_FONT
    ws.cell(row=r, column=3, value=label).font = SUB_FONT
    r += 1

# ---------- Sheet 2: Задачи ----------
ws2 = wb.create_sheet("Задачи")
cols2 = [("A",5),("B",16),("C",36),("D",46),("E",50),("F",54),("G",40),("H",26)]
for c,w in cols2: ws2.column_dimensions[c].width = w
head2 = ["№","Тип","Область","Что было / причина","Корень / детали","Что сделано","Файлы","Статус"]
for i,h in enumerate(head2, start=1):
    c = ws2.cell(row=1, column=i, value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=WRAP_CTR; c.border=BORDER
tasks = [
    ["1","Переработка (UX)","Чат — видео-кружки (Telegram-style)",
     "Запись «видеокружка» шла через белую Modal/Drawer-карточку, с in-app объяснением доступа и ручной красной кнопкой; не было паузы и прогресса загрузки",
     "Оверлей должен быть в portal(document.body) — иначе fixed-ребёнок клипается transform/overflow-hidden предком чата; авто-старт один раз за открытие",
     "Клик по камере сразу стартует запись; прозрачный полноэкранный оверлей (Telegram): таймер с десятыми + красная точка, «Отмена», пауза/резюме, «Отправить»; реальный MediaRecorder.pause()/resume() (таймер/кольцо замирают); оптимистичный пузырёк с кольцом прогресса и крестиком отмены загрузки",
     "useVideoNoteRecorder.ts, ChatVideoNoteRecorder.tsx (rewrite), ChatVideoNoteMessage.tsx, MessageBubble.tsx, ChatPage.tsx, useChatMediaUploadMutation.ts, useVideoNoteUploadStore.ts (new), chatVideoNoteRecorder.ts, chatHelpers.ts (isAbortError), locales ×15",
     "✅ Закоммичено 48c61aa5"],
    ["2","Переработка (UI)","Видеозвонок — Telegram-стиль полноэкранный оверлей",
     "Маленькая плавающая перетаскиваемая карточка звонка с maximize-переключателем",
     "Ранее отсутствовали ключи declineCall/acceptCall/endCall во всех локалях",
     "Полноэкранный fixed inset-0 оверлей: object-cover видео edge-to-edge, self-view PiP снизу-справа, докнутые кнопки (CallControlButton), центрированный аватар для аудио/звонка; «свернуть в угол» вместо maximize; role=dialog/aria-modal, фокус, Escape-сворачивание; +4 i18n-ключа ×15",
     "GlobalAudioCallModal.tsx (полная переработка return-блока, CallControlButton, captionLabel), locales ×15",
     "✅ Закоммичено 48c61aa5"],
    ["3","Фикс (bug)","Видеозвонок — ЗВОНЯЩИЙ не видел веб-камеру собеседника",
     "A звонит B, B принимает — A видит только свою камеру в PiP, видео собеседника не появляется; B видит оба потока",
     "Регрессия коммита 5a9235ab: удалён сигнал track.onunmute→setHasRemoteVideo(!track.muted); заменён на ненадёжные DOM-frame события, а track.onunmute был гейтнут el.videoWidth>0 (=0 при unmute → no-op) — не срабатывал у оферера",
     "Восстановлен track-mute сигнал в pc.ontrack (onunmute/onmute/onended); handleRemoteVideoFrames теперь только подтверждает true; добавлен offerInFlightRef (идемпотентность оффера — 3 некоординированных триггера не гонят 2 оффера)",
     "GlobalAudioCallModal.tsx (ontrack, createAndSendOffer + offerInFlightRef, cleanupCall), locales ×15",
     "✅ Закоммичено 48c61aa5; ⚠ live-раунд 2 абонентов"],
    ["4","Фикс (bug)","Видеозвонок — камера-OFF замораживала кадр; модалка терялась при reload",
     "Собеседник выключил камеру — его последний кадр «замерзал» на экране; при перезагрузке страницы модалка звонка не возвращалась, хотя звонок жив на сервере",
     "replaceTrack(null) держит трансивер sendrecv и часто НЕ шлёт mute-событие → hasRemoteVideo оставался true; состояние звонка — React-стейт, теряется при reload; нет эндпоинта «мой активный звонок»",
     "Watchdog живости кадров (requestVideoFrameCallback + 400мс интервал, порог ~1.2с, fallback по currentTime) очищает сцену за ~1.2с; персист id активного звонка + восстановление на WS-open (GET /calls/:id → hydrate/restoreMedia если ACTIVE/RINGING)",
     "GlobalAudioCallModal.tsx (FrameCallbackVideo, watchdog-эффект, ACTIVE_CALL_ID персист, syncCallsAfterReconnect)",
     "✅ Закоммичено a5d6b315; ⚠ media после reload best-effort"],
    ["5","Фикс (bug)","Ревью-фиксы звонка: cross-tab hijack, watchdog, phone/name",
     "Мульти-ракурсное ревью вскрыло 3 реальные регрессии этого дня",
     "(1) id звонка персистился в origin-global localStorage → вторая вкладка/юзер перехватывали чужой звонок; (2) watchdog выключал <video>, а потом сам себя ослеплял (if !el return); (3) /\\d/.test(query) слал любой запрос с цифрой в phone-поиск («Logistic 24» терялся)",
     "(1) sessionStorage + scope по userId ({userId,callId}); (2) hasRemoteVideoTrack держит <video> смонтированным; (3) общий isPhoneLikeQuery (цифра И phone-shape) в обоих call-site + тест",
     "GlobalAudioCallModal.tsx, CargoDriverOfferPanel.tsx, SendOfferToDriverModal.tsx, isPhoneLikeQuery.ts (new) + тест",
     "✅ Закоммичено a5d6b315"],
    ["6","Фикс (bug)","Груз — метка города содержала админ-описатели",
     "В table-/row-/card-видах и офферах from/to показывали «Провинция Хамгён-Намдо», «Кашкадаринска област», «Административный Эр-Рияд», полный адрес или сокращение «Toshkent Vil.»",
     "cityNameFromAddress не знал форм: провинция/провинции, «област»/«обл.», administrativ*/municipal*/округ/район; freeform-адрес геокодера не сводится к городу; region в прилагательной форме",
     "Расширены GEO_ADMIN_WORDS + GEO_QUALIFIER_RE (Safari-safe, без lookbehind); чистка орфанной «.»; summarizeCargoRoute отдаёт originCity/destinationCity = backend from_city_name/to_city_name (чистое локализованное имя) с fallback на стриппинг; полный адрес — в тултипе",
     "addressDisplay.ts, cargoListDisplay.ts, myCargosListCells.tsx, allCargosListCells.tsx + addressDisplay.test / cargoRouteSummary.test",
     "✅ Закоммичено a5d6b315, 37ceaf43"],
    ["7","Переработка (UI)","Груз — строковый вид (all-cargos): выравнивание сетки",
     "Мини-карты разной высоты по строкам; колонка «машины/статус/дата» дрейфовала; маршрут from/to плавал в вертикальном центре; после укорочения городов — большой зазор до груза",
     "@2xl-грид с items-stretch тянул колонки на переменную высоту строки; map/meta/route наследовали её; location-трек 190–240px был размером под полный адрес",
     "Карта self-start + фикс h-52 (208px) для обеих веток → одинаковые thumbnail'ы; meta justify-start self-start → top-align; route self-stretch → таймлайн растянут (откуда сверху / куда снизу); шаблон грида ретюнен (2fr/1fr) — груз ближе к локации, meta левее, цена справа",
     "CargoRouteRow.tsx (grid-cols template, map-ветки, route/meta-колонки)",
     "✅ Закоммичено 37ceaf43"],
    ["8","Фикс (bug)","Груз — «Отправить оффер»: поиск водителя/ДМ по телефону",
     "Селекты водителя и driver-manager искались только по имени; ввод телефона ничего не находил",
     "Серверный поиск шёл по name; клиентский фильтр по optionFilterProp=label (label = имя); телефон не нёсся на опциях",
     "Общий optionMatchesNameOrPhone (имя ИЛИ телефон, как есть и digits-only) как filterOption на обоих селектах; phone добавлен на обе option-формы; серверный запрос с цифрами уходит как phone-параметр",
     "CargoDriverOfferPanel.tsx (optionMatchesNameOrPhone, phone-маппинг, driverSearchParams)",
     "✅ Закоммичено a5d6b315; ⚠ live-проверка оффер-флоу"],
    ["9","Фикс (bug)","Карточка водителя — «Маълумот йўқ» для не пришедших полей",
     "Detail-вьюхи водителя всегда рисовали полный шаблон и ставили «No data» для отсутствующих полей — неотличимо от явного null (тикет SARB-TASK-FRONT)",
     "driver?.field даёт undefined и при отсутствии ключа, и при null; renderValue схлопывал оба в noData; поля рисовались безусловно",
     "Новый isDriverFieldPresent (через `key in driver`); presence-гейтед field()/readField рисуют ячейку только при наличии ключа; секции Documents/Vehicle/Activity + подблоки скрываются целиком; null всё ещё показывает «No data»",
     "driverFieldPresence.ts (new) + тест, DriverDetailsDrawer.tsx, DriverDetailsPage.tsx",
     "✅ Закоммичено a5d6b315; ⚠ live short-card vs full"],
    ["10","Функция","Админ — адаптив всей консоли (4K → 360px телефон)",
     "Требование: тело страницы никогда не скроллит горизонтально ни на одном брейкпоинте, модалки не «багованные», layout чистый (9-slice аудит: 39 находок; 3-lens ревью: 5 подтверждённых)",
     "Таблицы без scroll={{x}} (вытекали из карточки); ant-table-pagination — нескролящийся flex-ряд; Segmented (до 10 вкладок) в один ряд; Tailwind-override AntD был no-op (cssinjs unlayered бьёт @layer utilities)",
     "Shared-first: новый adminResponsive.css через ConfigProvider modal/drawer → каждый оверлей капит высоту + внутр. скролл + overflow-wrap:anywhere !important; пагинация reflow; Segmented в .admin-scroll-x; <main> overflow-x-hidden; таргетные фиксы (command-center scroll={{x}}, stat-grids grid-cols-1 sm:3, Push Space wrap, ExcelViewer min(500px,60vh), GeoLayerControls flex-wrap)",
     "adminResponsive.css (new), adminGridTable.css, AdminChrome.tsx, AdminAnalyticsHub.tsx, AdminPlatformOverview.tsx, AdminLayout.tsx, command-center/push/metrics/geo/users-*",
     "✅ Закоммичено 285fbfd6, 9b17a6b8; ⚠ live authenticated GUI"],
    ["11","Функция","Админ — тренд-график dashboard + карта водителя + call-debug",
     "Командному центру не хватало тренд-графика series; карточке водителя — живой мини-карты локации; двух-пирный звонок трудно диагностировать без инструментов",
     "series-поле бэкенда не контракт-locked (defensive parse); мини-карта Yandex как в all-cargos; call-debug должен быть off-by-default в проде",
     "adminDashboardSeries.ts (парсер бакетов day/week/month) + AdminLine.tsx (recharts, линия на метрику) в командном центре; DriverLocationCard/MiniMap (native Yandex + deep-link в GPS); callDebug.ts (opt-in localStorage['sarbon:call-debug']==='1', no-op иначе)",
     "adminDashboardSeries.ts (new)+тест, AdminLine.tsx (new), DriverLocationCard.tsx (new), DriverLocationMiniMap.tsx (new), callDebug.ts (new)+тест",
     "✅ Закоммичено a5d6b315, 285fbfd6"],
]
for ri,row in enumerate(tasks, start=2):
    for ci,val in enumerate(row, start=1):
        c = ws2.cell(row=ri, column=ci, value=val); c.font=CELL_FONT; c.alignment=WRAP_TOP; c.border=BORDER

# ---------- Sheet 3: Изменённые файлы ----------
ws3 = wb.create_sheet("Изменённые файлы")
cols3 = [("A",62),("B",26),("C",66),("D",26)]
for c,w in cols3: ws3.column_dimensions[c].width = w
head3 = ["Файл","Тип изменения","Что сделано","Коммит"]
for i,h in enumerate(head3, start=1):
    c = ws3.cell(row=1, column=i, value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=WRAP_CTR; c.border=BORDER
files = [
    ["src/components/shared/GlobalAudioCallModal.tsx","Логика/UI (звонки)","Каллер-веб-камера фикс, Telegram-редизайн, watchdog кадров, reload-восстановление, cross-tab фикс","48c61aa5, a5d6b315, 285fbfd6, 9b17a6b8"],
    ["src/hooks/useVideoNoteRecorder.ts","Хук (кружки)","Состояние 'paused', pause/resume, tick-driven авто-стоп, swapCameraDuringActiveRecording","48c61aa5"],
    ["src/pages/chat/components/ChatVideoNoteRecorder.tsx","UI (кружки, rewrite)","createPortal-оверлей, авто-старт записи, Telegram-бар, on-dark accent","48c61aa5"],
    ["src/pages/chat/components/ChatVideoNoteMessage.tsx","UI (кружки)","Кольцо прогресса загрузки + X-отмена (подписка по uploadId)","48c61aa5"],
    ["src/stores/useVideoNoteUploadStore.ts","Стор (новый)","Zustand: прогресс загрузки + AbortController'ы по temp-id","48c61aa5"],
    ["src/pages/chat/hooks/useChatMediaUploadMutation.ts","Хук (чат)","Проброс signal + onUploadProgress в media-POST","48c61aa5"],
    ["src/pages/chat/utils/chatVideoNoteRecorder.ts","Утилита (кружки)","formatVideoNoteElapsed(ms) → m:ss,t","48c61aa5"],
    ["src/utils/chatHelpers.ts","Утилита (чат)","isAbortError-guard (подавляет тост отменённой загрузки)","48c61aa5"],
    ["src/utils/addressDisplay.ts","Утилита (адрес)","cityNameFromAddress: дроп провинция/област/administrativ*/municipal*/округ/район; чистка орфанной пунктуации","a5d6b315, 37ceaf43"],
    ["src/utils/cargoListDisplay.ts","Утилита (груз)","CargoRouteSummary.originCity/destinationCity = backend from/to_city_name","a5d6b315"],
    ["src/pages/cargos/all-cargos/components/CargoRouteRow.tsx","UI (груз, row)","Карта self-start/фикс-высота, meta top-align, route self-stretch, ретюн грида","37ceaf43"],
    ["src/pages/cargos/my-cargos/utils/myCargosListCells.tsx","UI (груз)","Метка города → r.originCity/destinationCity","a5d6b315"],
    ["src/pages/cargos/all-cargos/utils/allCargosListCells.tsx","UI (груз)","Метка города, дроп fullAddress-ветки","a5d6b315"],
    ["src/pages/cargos/components/CargoDriverOfferPanel.tsx","UI/логика (оффер)","optionMatchesNameOrPhone, phone на опциях, phone-vs-name серверный поиск","a5d6b315"],
    ["src/pages/drivers/components/SendOfferToDriverModal.tsx","UI (оффер)","Общий isPhoneLikeQuery","a5d6b315"],
    ["src/utils/isPhoneLikeQuery.ts","Утилита (новая)","Строгий предикат phone-запроса (цифра И phone-shape) (+тест)","a5d6b315"],
    ["src/utils/driverFieldPresence.ts","Утилита (новая)","isDriverFieldPresent через `key in driver` (+тест)","a5d6b315"],
    ["src/pages/drivers/components/DriverDetailsDrawer.tsx","UI (водитель)","Presence-гейтинг ячеек/секций; скрытие отсутствующих полей","a5d6b315"],
    ["src/pages/drivers/DriverDetailsPage.tsx","UI (водитель)","Presence-гейтинг ячеек/секций (полная страница)","a5d6b315"],
    ["src/pages/drivers/components/DriverLocationCard.tsx","UI (новый, водитель)","Карточка live-GPS: Yandex мини-карта + deep-link в GPS-трекинг","a5d6b315"],
    ["src/pages/drivers/components/DriverLocationMiniMap.tsx","UI (новый, водитель)","Нативная Yandex мини-карта, центрированная на водителе","a5d6b315"],
    ["src/components/shared/admin/adminResponsive.css","CSS (новый, админ)","Оверлеи капят высоту + внутр. скролл + overflow-wrap:anywhere; пагинация reflow; Segmented scroll-x","285fbfd6, 9b17a6b8"],
    ["src/layouts/admin-layout/AdminLayout.tsx","Layout (админ)","ConfigProvider modal/drawer конфиг; <main> overflow-x-hidden backstop","285fbfd6, 9b17a6b8"],
    ["src/pages/admin/command-center/AdminCommandCenterPage.tsx","UI (админ)","scroll={{x}} на 2 таблицах; тренд-график series","a5d6b315, 285fbfd6"],
    ["src/utils/adminDashboardSeries.ts","Утилита (новая, админ)","Defensive-парсер series (bucket/timeKey/metricKeys) (+тест)","a5d6b315"],
    ["src/pages/admin/command-center/components/charts/AdminLine.tsx","UI (новый, чарт)","Recharts тайм-серия: линия на метрику","a5d6b315"],
    ["src/utils/callDebug.ts","Утилита (новая)","Opt-in диагностика звонка (localStorage['sarbon:call-debug']) (+тест)","285fbfd6"],
    ["src/utils/cargoGeo.ts · cargoGeoFacets · citySelectOptions · Ci­ty/CountrySelect","Утилита/UI","Уточнения geo-фасетов, опций города/страны","285fbfd6, 9b17a6b8"],
    ["public/locales/*.json (×15)","Локализация","Ключи звонка (decline/accept/end/isCallingYou), кружков (cancel/pause/resume)","48c61aa5, a5d6b315"],
    ["bug.md","Документация","11 записей 04.07 (10:40 … 16:40)","верх файла"],
]
for ri,row in enumerate(files, start=2):
    for ci,val in enumerate(row, start=1):
        c = ws3.cell(row=ri, column=ci, value=val); c.font=CELL_FONT; c.alignment=WRAP_TOP; c.border=BORDER

# ---------- Sheet 4: QA - Статус ----------
ws4 = wb.create_sheet("QA - Статус")
cols4 = [("A",46),("B",30),("C",68)]
for c,w in cols4: ws4.column_dimensions[c].width = w
head4 = ["Проверка","Результат","Примечание"]
for i,h in enumerate(head4, start=1):
    c = ws4.cell(row=1, column=i, value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=WRAP_CTR; c.border=BORDER
qa = [
    ["pnpm build (tsc -b + vite)","✅ exit 0","Полная сборка проекта — единственная воротная проверка CI"],
    ["eslint --max-warnings 0","✅ 0 предупреждений","На всех затронутых файлах (pre-commit)"],
    ["vitest (полный прогон)","✅ 718 / 718","Новые/расширенные: driverFieldPresence, isPhoneLikeQuery, adminDashboardSeries, callDebug, cargoGeo, addressDisplay, cargoRouteSummary, chatVideoNoteRecorder, webrtcTuning, citySelectOptions"],
    ["Cascade !important (adminResponsive)","✅ подтверждено","Проверено в скомпилированном dist/assets/*.css (Tailwind vs AntD cssinjs)"],
    ["Метка города — стриппинг","✅ юнит-тесты","addressDisplay.test 17/17; ложноположительный guard «Облучье» сохранён"],
    ["verify — видеозвонок (caller видит callee)","⚠ требует QA","P2P между 2 абонентами A↔B под авторизацией; в среде отчёта недоступно"],
    ["verify — камера-off / reload-восстановление","⚠ требует QA","Нужен reload и вкл/выкл камеры на живом звонке"],
    ["verify — видео-кружки (пауза/загрузка)","⚠ требует QA","Гейтится камерой/мик. и HTTPS; нужен клик на устройстве с камерой"],
    ["verify — адаптив админ-консоли","⚠ требует QA","capture-admin-responsive.mjs (4K/laptop/tablet/360px) на dev-сервере против staging; проверяет overflow=0"],
    ["verify — строковый вид / карточка водителя","⚠ требует QA","capture-all-cargos-row.mjs; short-card vs full-profile ответ"],
    ["Cross-tab / cross-user звонок","✅ фикс (код)","sessionStorage + scope по userId; live мульти-таб раунд — на staging"],
]
for ri,row in enumerate(qa, start=2):
    for ci,val in enumerate(row, start=1):
        c = ws4.cell(row=ri, column=ci, value=val); c.font=CELL_FONT; c.alignment=WRAP_TOP; c.border=BORDER

for w in (ws2, ws3, ws4):
    w.freeze_panes = "A2"

wb.save(OUT)
print("wrote", OUT)
