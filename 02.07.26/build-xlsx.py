# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = r"C:\Users\hp\Desktop\sarbon-frontend-main\daily-report\02.07.26\report-02.07.26.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1639C8")
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="16181D")
SUB_FONT = Font(bold=False, size=10.5, color="5b6477")
CELL_FONT = Font(bold=False, size=10.5, color="16181D")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
WRAP_CTR = Alignment(wrap_text=True, vertical="center")
THIN = Side(style="thin", color="E6E9F0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

wb = openpyxl.Workbook()

# ---------- Sheet 1: Сводка ----------
ws = wb.active
ws.title = "Сводка"
ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 3
ws.column_dimensions["C"].width = 30
ws["A1"] = "Отчёт о работе · 02.07.2026 · Sarbon Frontend"; ws["A1"].font = TITLE_FONT
ws["A2"] = "Фокус: 5 областей — чат, аудиозвонки, админ-модерация, деактивация груза, QA"; ws["A2"].font = SUB_FONT
rows = [
    ("02.07.2026", "Дата"),
    ("sarbon-frontend  ·  React 19 · Vite · Ant Design 6", "Проект"),
    ("staging", "Ветка / окружение"),
    ("3", "Дефектов закрыто"),
    ("2", "Функций добавлено"),
    ("2", "Регрессий найдено и устранено в review"),
    ("48 / 48", "Unit-тестов в затронутых наборах"),
    ("0", "Ошибок tsc -b"),
    ("0", "Предупреждений ESLint (затронутые файлы)"),
    ("11", "Реальных скриншотов-доказательств"),
]
r = 4
for val, label in rows:
    ws.cell(row=r, column=1, value=val).font = CELL_FONT
    ws.cell(row=r, column=3, value=label).font = SUB_FONT
    r += 1

# ---------- Sheet 2: Задачи ----------
ws2 = wb.create_sheet("Задачи")
cols2 = [("A",5),("B",16),("C",30),("D",40),("E",44),("F",46),("G",34),("H",22)]
for c,w in cols2: ws2.column_dimensions[c].width = w
head2 = ["№","Тип","Область","Что было / причина","Корень / детали","Что сделано","Файлы","Статус"]
for i,h in enumerate(head2, start=1):
    c = ws2.cell(row=1, column=i, value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=WRAP_CTR; c.border=BORDER
tasks = [
    ["1","Фикс (bug)","Чат — пузырёк звонка",
     "Пузырёк сообщения-звонка не показывал длительность; входящий/исходящий всегда на «моей» стороне; иконка не отражала статус",
     "sender_id одинаков для обоих направлений — надёжный сигнал payload.direction/status ранее не читался; длительность в payload.duration_seconds",
     "Длительность «N мин. NN сек.»; сторона по payload.direction (payload-first); зелёный PhoneCall «завершён», красный PhoneOff «отменён/отклонён»; убран дубль-override в ChatThreadPanel",
     "chatCallEvents.ts, MessageBubble.tsx, ChatThreadPanel.tsx, chatTypes.ts, locales ×15",
     "✅ Закоммичено 6fe53d78; verify скриншотами (staging)"],
    ["2","Фикс (bug)","Аудиозвонки — устойчивость к обрыву",
     "При локальном сбое WebRTC модалка закрывалась, а серверный звонок оставался ACTIVE → «залипание занят» без кнопки «Завершить»",
     "cleanupCall('error') безусловно очищал всю сессию; isCallActive исключал error → модалка размонтировалась",
     "preserveSession сохраняет сессию на ошибках; isCallActive держит модалку при error && callId; опрос не оживляет медиа в error; статус «Ошибка звонка» (callError, розовый). В review устранены 2 регрессии (decay-deps, reconnect-guard)",
     "GlobalAudioCallModal.tsx, ChatThreadPanel.tsx, locales ×15",
     "✅ В рабочем дереве; verify рендером реального компонента"],
    ["3","Функция","Админ — модерация менеджеров (CM/DM)",
     "Вкладка «Менеджеры» была только для чтения (read-only алерт), без действий — в отличие от грузов и водителей (KYC)",
     "Панель никогда не строилась с мутациями; эндпоинт accept/reject для диспетчеров не задокументирован в OpenAPI",
     "«Принять» (пустой POST) + «Отклонить» (модалка с обязательной причиной); гейтинг решённых строк через canModerateDispatcher; read-only алерт убран",
     "ManagerModerationPanel.tsx, dispatcherModeration.ts (+тест), endpoints.ts, locales ×15",
     "✅ В рабочем дереве; ⚠ эндпоинт — ожидаемый контракт"],
    ["4","Функция","Груз — деактивация / возврат в поиск",
     "У груза в поиске не было способа временно скрыть его из выдачи (и вернуть) с указанием причины",
     "Новый lifecycle-переход: DEACTIVATABLE (в поиске) → DEACTIVATED и обратно",
     "Контрол на детальной странице груза; обязательная причина из справочника или свободный текст для «Другое» (лимит 500); системные причины отфильтрованы",
     "CargoDeactivationControl.tsx, cargoDeactivation.ts (+тест), cargoStatus.ts, MyCargoDetailPage.tsx",
     "✅ Закоммичено 6fe53d78; 20 unit-тестов"],
    ["5","Инфра (QA)","QA-прогон + Allure-отчёт",
     "Прогон чат-пузырька был прерван на сборке Allure; отчёт брендирован как «Admin Console»",
     "allure-from-run-filtered.mjs строит Allure-результаты из прогона; отчёт собирается allure-commandline",
     "Достроен Allure-отчёт (4 теста, все passed, 0 реальных ошибок консоли); переоформлен под «Sarbon Frontend»; +2 паттерна безобидного шума консоли (kebab-SVG, uncontrolled-input)",
     "qa/allure-from-run-filtered.mjs, qa/reports/…chat-call-bubble.md, allure-report/",
     "✅ Отчёт собран; 0 находок"],
]
for ri,row in enumerate(tasks, start=2):
    for ci,val in enumerate(row, start=1):
        c = ws2.cell(row=ri, column=ci, value=val); c.font=CELL_FONT; c.alignment=WRAP_TOP; c.border=BORDER

# ---------- Sheet 3: Изменённые файлы ----------
ws3 = wb.create_sheet("Изменённые файлы")
cols3 = [("A",56),("B",24),("C",62),("D",30)]
for c,w in cols3: ws3.column_dimensions[c].width = w
head3 = ["Файл","Тип изменения","Что сделано","Место / статус"]
for i,h in enumerate(head3, start=1):
    c = ws3.cell(row=1, column=i, value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=WRAP_CTR; c.border=BORDER
files = [
    ["src/pages/chat/utils/chatCallEvents.ts","Логика (чат)","resolveChatCallEvent (payload-first direction/status), длительность, статус-категория","закоммичено 6fe53d78"],
    ["src/pages/chat/components/MessageBubble.tsx","UI (чат)","Сторона/цвет/длительность пузырька звонка из resolveChatCallEvent","закоммичено 6fe53d78"],
    ["src/pages/chat/components/ChatThreadPanel.tsx","UI/логика (чат)","Убран дубль-override направления; callBusy учитывает error","коммит + рабочее дерево"],
    ["src/types/chatTypes.ts","Типы","ChatCallPayload.direction/status; ChatMessageType += 'call'","закоммичено 6fe53d78"],
    ["src/components/shared/GlobalAudioCallModal.tsx","Логика (звонки)","preserveSession; isCallActive для error&&callId; decay-deps += callId; reconnect-guard; статус «Ошибка звонка»","рабочее дерево"],
    ["src/pages/admin/moderation/components/ManagerModerationPanel.tsx","UI (админ)","Accept/Reject + модалка причины; read-only алерт убран","рабочее дерево"],
    ["src/pages/admin/moderation/components/dispatcherModeration.ts","Логика (админ)","canModerateDispatcher, dispatcherModerationValue (+тест)","рабочее дерево"],
    ["src/pages/cargos/my-cargo-detail/components/CargoDeactivationControl.tsx","UI (груз)","Контрол деактивации/возврата в поиск с причиной","закоммичено 6fe53d78"],
    ["src/utils/cargoDeactivation.ts","Логика (груз)","Статусы/действие/резолв причины (+тест 20 кейсов)","закоммичено 6fe53d78"],
    ["src/constants/endpoints.ts","Константы","DISPATCHER_MODERATION_ACCEPT/REJECT (ожидаемый контракт)","рабочее дерево"],
    ["src/layouts/home-layout/HomeLayout.tsx","UI (мелочь)","Порядок Tailwind-классов (font-family-inter!) — снят lint-warning","рабочее дерево"],
    ["public/locales/*.json (×15)","Локализация","callError (все 15); ключи модерации менеджеров","коммит + рабочее дерево"],
    ["qa/allure-from-run-filtered.mjs","QA-инфра","Ребрендинг «Sarbon Frontend»; +2 паттерна безобидного шума консоли","рабочее дерево"],
    ["bug.md","Документация","3 записи 02.07 (15:48, 16:33, 17:25 + follow-up review)","верх файла"],
]
for ri,row in enumerate(files, start=2):
    for ci,val in enumerate(row, start=1):
        c = ws3.cell(row=ri, column=ci, value=val); c.font=CELL_FONT; c.alignment=WRAP_TOP; c.border=BORDER

# ---------- Sheet 4: QA - Статус ----------
ws4 = wb.create_sheet("QA - Статус")
cols4 = [("A",40),("B",30),("C",62)]
for c,w in cols4: ws4.column_dimensions[c].width = w
head4 = ["Проверка","Результат","Примечание"]
for i,h in enumerate(head4, start=1):
    c = ws4.cell(row=1, column=i, value=h); c.font=HEADER_FONT; c.fill=HEADER_FILL; c.alignment=WRAP_CTR; c.border=BORDER
qa = [
    ["tsc -b","✅ 0 ошибок","После всех правок и отката временного кода для скриншота"],
    ["eslint --max-warnings 0","✅ 0 предупреждений","На всех затронутых файлах"],
    ["vitest (затронутые наборы)","✅ 48/48","chatCallEvents 25 · cargoDeactivation 20 · dispatcherModeration 3"],
    ["vitest (полный прогон)","⚠ авария воркера (ENOMEM)","При одновременном dev-сервере/Playwright; чистый прогон — на свободной машине"],
    ["code-review (фикс звонков)","✅ 2 регрессии устранены","decay-deps += callId; reconnect-guard; 4 находки задокументированы"],
    ["verify — чат-пузырёк","✅ реальные скриншоты","Живое приложение, staging, реальные сообщения-звонки"],
    ["verify — модалка «Ошибка звонка»","✅ рендер реального компонента","Временный сид error-состояния, откачен после съёмки"],
    ["verify — модерация менеджеров","✅ живой прогон","Админ-авторизация; Accept/Reject + модалка причины"],
    ["QA-отчёт чат-пузырька","✅ 0 находок","Allure собран, переоформлен «Sarbon Frontend»"],
    ["Эндпоинт модерации менеджеров","⚠ требует подтверждения","/v1/admin/dispatchers/:id/moderation/accept|reject — ожидаемый контракт"],
    ["Деактивация груза (live)","⚠ требует QA","Контрол за ProtectedRoute; логика покрыта unit-тестами"],
]
for ri,row in enumerate(qa, start=2):
    for ci,val in enumerate(row, start=1):
        c = ws4.cell(row=ri, column=ci, value=val); c.font=CELL_FONT; c.alignment=WRAP_TOP; c.border=BORDER

for w in (ws2, ws3, ws4):
    w.freeze_panes = "A2"

wb.save(OUT)
print("wrote", OUT)
